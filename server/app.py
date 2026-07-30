"""
Cruce Stock Farmacias — Flask + SQLite
Backend: auth JWT, cruces, lineas, estados cadete, analytics, descarga Excel.
"""

import io
import json
import os
import re
import secrets
import sys
import threading
import time
import uuid

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timezone, timedelta
from functools import wraps
from pathlib import Path

import jwt
import yaml
from flask import Flask, Response, g, jsonify, request, send_from_directory, send_file, stream_with_context
from werkzeug.security import check_password_hash, generate_password_hash
from rapidfuzz import process as fuzz_process, fuzz

BASE_DIR    = Path(__file__).parent.parent
SERVER_DIR  = Path(__file__).parent
DB_PATH     = Path(os.environ.get('CRUCE_DB_PATH',      str(SERVER_DIR / 'cruce.db')))
UPLOADS_DIR = Path(os.environ.get('CRUCE_UPLOADS_DIR',  str(SERVER_DIR / 'uploads')))
CONFIG_PATH = Path(os.environ.get('CRUCE_CONFIG_PATH',  str(BASE_DIR / 'cruce_stock' / 'config.yaml')))

SECRET_KEY = os.environ.get('CRUCE_SECRET', 'dev-secret-change-in-prod')
TOKEN_TTL  = int(os.environ.get('CRUCE_TOKEN_TTL', 86400))

# Agregar src al path para importar los módulos de procesamiento
sys.path.insert(0, str(BASE_DIR / 'cruce_stock'))

app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path='')

UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# Si CONFIG_PATH apunta a un volumen externo y todavía no existe,
# copiamos el default del repositorio para no arrancar sin configuración
_config_default = BASE_DIR / 'cruce_stock' / 'config.yaml'
if not CONFIG_PATH.exists() and CONFIG_PATH != _config_default and _config_default.exists():
    import shutil
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(_config_default, CONFIG_PATH)

# ── DB ─────────────────────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        import sqlite3
        g.db = sqlite3.connect(str(DB_PATH), detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA journal_mode=WAL')
        g.db.execute('PRAGMA foreign_keys=ON')
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db:
        db.close()

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario      TEXT    UNIQUE NOT NULL,
            password_hash TEXT   NOT NULL,
            nombre       TEXT,
            rol          TEXT    DEFAULT 'operador',
            created_at   TEXT    DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS cruces (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha           TEXT    NOT NULL,
            archivo_pedidos TEXT,
            archivo_stock   TEXT,
            total_lineas    INTEGER DEFAULT 0,
            con_cobertura   INTEGER DEFAULT 0,
            sin_cobertura   INTEGER DEFAULT 0,
            pct_cobertura   REAL    DEFAULT 0,
            excel_path      TEXT,
            usuario_id      INTEGER REFERENCES usuarios(id),
            created_at      TEXT    DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS lineas (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            cruce_id         INTEGER NOT NULL REFERENCES cruces(id) ON DELETE CASCADE,
            nro_pedido       TEXT,
            gtin             TEXT,
            sku              TEXT,
            producto         TEXT,
            marca            TEXT,
            unidades         INTEGER DEFAULT 1,
            nodo_asignado    TEXT,
            zona             TEXT,
            tier             INTEGER DEFAULT 2,
            stock_disponible INTEGER DEFAULT 0,
            alternativas     TEXT    DEFAULT '[]',
            estado           TEXT    DEFAULT 'PENDIENTE',
            notas            TEXT    DEFAULT '',
            updated_at       TEXT
        );

        CREATE TABLE IF NOT EXISTS eventos_estado (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            linea_id        INTEGER NOT NULL REFERENCES lineas(id) ON DELETE CASCADE,
            estado_anterior TEXT,
            estado_nuevo    TEXT,
            motivo          TEXT,
            origen          TEXT    DEFAULT 'cadete',
            created_at      TEXT    DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_lineas_cruce    ON lineas(cruce_id);
        CREATE INDEX IF NOT EXISTS idx_lineas_estado   ON lineas(estado);
        CREATE INDEX IF NOT EXISTS idx_eventos_linea   ON eventos_estado(linea_id);

        CREATE TABLE IF NOT EXISTS jobs (
            id         TEXT PRIMARY KEY,
            status     TEXT DEFAULT 'pending',
            step       TEXT DEFAULT 'En cola...',
            cruce_id   INTEGER,
            resumen    TEXT,
            error      TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)

    # Migraciones para versiones anteriores de la DB
    for col_sql in [
        "ALTER TABLE lineas ADD COLUMN alternativa_nodo TEXT DEFAULT ''",
        "ALTER TABLE lineas ADD COLUMN punto_retiro TEXT DEFAULT ''",
        "ALTER TABLE entregas_habilitadas ADD COLUMN nro_pedidos TEXT DEFAULT '[]'",
        "ALTER TABLE eventos_estado ADD COLUMN nodo_asignado TEXT",
        "ALTER TABLE lineas ADD COLUMN cliente_nombre TEXT DEFAULT ''",
        "ALTER TABLE lineas ADD COLUMN cliente_dni TEXT DEFAULT ''",
        "ALTER TABLE lineas ADD COLUMN cliente_telefono TEXT DEFAULT ''",
    ]:
        try:
            db.execute(col_sql)
            db.commit()
        except Exception:
            pass

    # Tabla de incidencias operativas (pedidos perdidos, reclamos, demoras, etc.)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS incidencias_operativas (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo        TEXT NOT NULL DEFAULT 'otro',
            nro_pedido  TEXT DEFAULT '',
            sucursal    TEXT DEFAULT '',
            descripcion TEXT NOT NULL,
            estado      TEXT DEFAULT 'abierta',
            resolucion  TEXT DEFAULT '',
            creado_por  TEXT,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)

    # Teléfonos de nodos/farmacias
    db.executescript("""
        CREATE TABLE IF NOT EXISTS nodos_config (
            nombre   TEXT PRIMARY KEY,
            telefono TEXT DEFAULT ''
        );
    """)

    # Tablas de entregas
    db.executescript("""
        CREATE TABLE IF NOT EXISTS entregas_confirmadas (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            cruce_id     INTEGER NOT NULL REFERENCES cruces(id) ON DELETE CASCADE,
            punto_retiro TEXT    NOT NULL,
            receptor     TEXT    NOT NULL,
            firma_base64 TEXT,
            entregado_at TEXT    DEFAULT CURRENT_TIMESTAMP,
            usuario_id   INTEGER REFERENCES usuarios(id)
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_entrega_unica
            ON entregas_confirmadas(cruce_id, punto_retiro);

        CREATE TABLE IF NOT EXISTS entregas_habilitadas (
            cruce_id     INTEGER NOT NULL REFERENCES cruces(id) ON DELETE CASCADE,
            punto_retiro TEXT    NOT NULL,
            nro_pedidos  TEXT    DEFAULT '[]',
            PRIMARY KEY (cruce_id, punto_retiro)
        );

        CREATE TABLE IF NOT EXISTS jornadas (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha      TEXT    NOT NULL,
            cadete_id  INTEGER NOT NULL REFERENCES usuarios(id),
            cruce_id   INTEGER NOT NULL REFERENCES cruces(id),
            puntos     TEXT    NOT NULL DEFAULT '[]',
            created_at TEXT    DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(fecha, cadete_id)
        );

        CREATE TABLE IF NOT EXISTS puntos_retiro_tokens (
            punto_retiro TEXT PRIMARY KEY,
            token        TEXT UNIQUE NOT NULL,
            created_at   TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS pedidos_retirados (
            punto_retiro TEXT NOT NULL,
            nro_pedido   TEXT NOT NULL,
            retirado_at  TEXT DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (punto_retiro, nro_pedido)
        );
    """)

    # Usuario admin por defecto si no existe ninguno
    existing = db.execute('SELECT COUNT(*) FROM usuarios').fetchone()[0]
    if existing == 0:
        db.execute(
            'INSERT OR IGNORE INTO usuarios (usuario, password_hash, nombre, rol) VALUES (?, ?, ?, ?)',
            ('admin', generate_password_hash('admin123'), 'Administrador', 'admin')
        )
        db.execute(
            'INSERT OR IGNORE INTO usuarios (usuario, password_hash, nombre, rol) VALUES (?, ?, ?, ?)',
            ('cadete', generate_password_hash('cadete123'), 'Cadete', 'cadete')
        )
    db.commit()

# ── Auth ────────────────────────────────────────────────────────────────────────

def _make_token(user_id, usuario, rol):
    payload = {
        'sub': str(user_id),
        'usr': usuario,
        'rol': rol,
        'exp': datetime.now(timezone.utc) + timedelta(seconds=TOKEN_TTL),
        'iat': datetime.now(timezone.utc),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

def _decode_token(token):
    return jwt.decode(token, SECRET_KEY, algorithms=['HS256'])

def _match_telefono(nombre: str, nodos_map: dict) -> str:
    """Fuzzy-match nombre contra nodos_config y devuelve el teléfono o ''."""
    if not nodos_map:
        return ''
    if nombre in nodos_map:
        return nodos_map[nombre]
    result = fuzz_process.extractOne(
        nombre, nodos_map.keys(),
        scorer=fuzz.token_sort_ratio,
        score_cutoff=75,
    )
    return nodos_map[result[0]] if result else ''

def require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        header = request.headers.get('Authorization', '')
        token_str = header[7:] if header.startswith('Bearer ') else request.args.get('token', '')
        if not token_str:
            return jsonify({'ok': False, 'motivo': 'sin_token'}), 401
        try:
            payload = _decode_token(token_str)
        except jwt.ExpiredSignatureError:
            return jsonify({'ok': False, 'motivo': 'token_vencido'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'ok': False, 'motivo': 'token_invalido'}), 401
        g.user_id  = int(payload['sub'])
        g.usuario  = payload['usr']
        g.rol      = payload['rol']
        return f(*args, **kwargs)
    return wrapper

def require_admin(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if getattr(g, 'rol', None) != 'admin':
            return jsonify({'ok': False, 'motivo': 'sin_permiso'}), 403
        return f(*args, **kwargs)
    return require_auth(wrapper)

# ── Rutas auth ──────────────────────────────────────────────────────────────────


@app.post('/api/auth/login')
def auth_login():
    data   = request.get_json(silent=True) or {}
    usuario = data.get('usuario', '').strip()
    password = data.get('password', '')
    if not usuario or not password:
        return jsonify({'ok': False, 'motivo': 'campos_requeridos'}), 400

    db  = get_db()
    row = db.execute('SELECT * FROM usuarios WHERE usuario = ?', (usuario,)).fetchone()
    if not row or not check_password_hash(row['password_hash'], password):
        return jsonify({'ok': False, 'motivo': 'credenciales_invalidas'}), 401

    token = _make_token(row['id'], row['usuario'], row['rol'])
    return jsonify({
        'ok': True,
        'token': token,
        'session': {
            'id':      row['id'],
            'usuario': row['usuario'],
            'nombre':  row['nombre'],
            'rol':     row['rol'],
        },
    })

@app.post('/api/auth/refresh')
@require_auth
def auth_refresh():
    db  = get_db()
    row = db.execute('SELECT * FROM usuarios WHERE id = ?', (g.user_id,)).fetchone()
    if not row:
        return jsonify({'ok': False}), 401
    token = _make_token(row['id'], row['usuario'], row['rol'])
    return jsonify({'ok': True, 'token': token})

@app.post('/api/auth/logout')
@require_auth
def auth_logout():
    return jsonify({'ok': True})

# ── Rutas cruces ────────────────────────────────────────────────────────────────

@app.get('/api/cruces')
@require_auth
def cruces_list():
    db   = get_db()
    rows = db.execute(
        '''SELECT c.*, u.nombre as usuario_nombre
           FROM cruces c LEFT JOIN usuarios u ON c.usuario_id = u.id
           ORDER BY c.created_at DESC LIMIT 50'''
    ).fetchall()
    return jsonify({'ok': True, 'cruces': [dict(r) for r in rows]})

@app.get('/api/cruces/<int:cruce_id>')
@require_auth
def cruce_detail(cruce_id):
    db  = get_db()
    row = db.execute('SELECT * FROM cruces WHERE id = ?', (cruce_id,)).fetchone()
    if not row:
        return jsonify({'ok': False, 'motivo': 'no_encontrado'}), 404
    lineas = db.execute(
        'SELECT * FROM lineas WHERE cruce_id = ? ORDER BY nro_pedido, nodo_asignado',
        (cruce_id,)
    ).fetchall()
    return jsonify({
        'ok':    True,
        'cruce': dict(row),
        'lineas': [dict(l) for l in lineas],
    })

@app.delete('/api/cruces/<int:cruce_id>')
@require_admin
def cruce_delete(cruce_id):
    db = get_db()
    if not db.execute('SELECT id FROM cruces WHERE id = ?', (cruce_id,)).fetchone():
        return jsonify({'ok': False, 'motivo': 'no_encontrado'}), 404
    # jornadas no tiene ON DELETE CASCADE, hay que borrar antes
    db.execute('DELETE FROM jornadas WHERE cruce_id = ?', (cruce_id,))
    db.execute('DELETE FROM cruces WHERE id = ?', (cruce_id,))
    db.commit()
    return jsonify({'ok': True})

# ── Jobs (procesamiento asíncrono) ──────────────────────────────────────────────

def _job_update(job_id, **kwargs):
    """Actualiza un job en la DB desde cualquier hilo (sin contexto Flask)."""
    import sqlite3 as _sqlite3
    kwargs['updated_at'] = datetime.now(timezone.utc).isoformat()
    sets = ', '.join(f'{k} = ?' for k in kwargs)
    conn = _sqlite3.connect(str(DB_PATH))
    try:
        conn.execute(f'UPDATE jobs SET {sets} WHERE id = ?', list(kwargs.values()) + [job_id])
        conn.commit()
    finally:
        conn.close()

def _run_job(job_id, path_pedidos, path_stock, run_dir, nombre_pedidos, nombre_stock, usuario_id):
    """Corre en un hilo separado. Procesa el cruce y guarda en la DB."""
    def upd(step):
        _job_update(job_id, step=step)

    try:
        _job_update(job_id, status='running')
        upd('Cargando y validando archivos...')
        resultado = _procesar_cruce(path_pedidos, path_stock, run_dir)

        upd('Guardando resultados en la base de datos...')
        with app.app_context():
            db = get_db()
            cur = db.execute(
                '''INSERT INTO cruces
                   (fecha, archivo_pedidos, archivo_stock, total_lineas, con_cobertura,
                    sin_cobertura, pct_cobertura, excel_path, usuario_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (
                    datetime.now(timezone.utc).isoformat(),
                    nombre_pedidos, nombre_stock,
                    resultado['total'], resultado['con_cobertura'],
                    resultado['sin_cobertura'], resultado['pct_cobertura'],
                    resultado.get('excel_path'), usuario_id,
                )
            )
            cruce_id = cur.lastrowid
            for linea in resultado['lineas']:
                db.execute(
                    '''INSERT INTO lineas
                       (cruce_id, nro_pedido, gtin, sku, producto, marca, unidades,
                        nodo_asignado, zona, tier, stock_disponible, alternativas, estado, punto_retiro,
                        cliente_nombre, cliente_dni, cliente_telefono)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        cruce_id,
                        linea.get('nro_pedido'), linea.get('gtin'), linea.get('sku'),
                        linea.get('producto'), linea.get('marca'), linea.get('unidades', 1),
                        linea.get('nodo_asignado'), linea.get('zona'), linea.get('tier', 2),
                        linea.get('stock_disponible', 0),
                        json.dumps(linea.get('alternativas', []), ensure_ascii=False),
                        'SIN_COBERTURA' if not linea.get('nodo_asignado') else 'PENDIENTE',
                        linea.get('punto_retiro', ''),
                        linea.get('cliente_nombre', ''), linea.get('cliente_dni', ''), linea.get('cliente_telefono', ''),
                    )
                )
            db.commit()

        _job_update(job_id, status='done', cruce_id=cruce_id,
                    resumen=json.dumps(resultado, ensure_ascii=False))

    except Exception as e:
        _job_update(job_id, status='error', error=str(e))


@app.post('/api/cruces')
@require_auth
def cruce_nuevo():
    if 'pedidos' not in request.files or 'stock' not in request.files:
        return jsonify({'ok': False, 'motivo': 'archivos_requeridos'}), 400

    f_pedidos = request.files['pedidos']
    f_stock   = request.files['stock']

    run_id  = uuid.uuid4().hex
    run_dir = UPLOADS_DIR / run_id
    run_dir.mkdir()

    path_pedidos = run_dir / f_pedidos.filename
    path_stock   = run_dir / f_stock.filename
    f_pedidos.save(str(path_pedidos))
    f_stock.save(str(path_stock))

    job_id = uuid.uuid4().hex
    db = get_db()
    db.execute('INSERT INTO jobs (id, status, step) VALUES (?, ?, ?)',
               (job_id, 'pending', 'En cola...'))
    db.commit()

    t = threading.Thread(
        target=_run_job,
        args=(job_id, str(path_pedidos), str(path_stock), str(run_dir),
              f_pedidos.filename, f_stock.filename, g.user_id),
        daemon=True,
    )
    t.start()

    return jsonify({'ok': True, 'job_id': job_id})


@app.get('/api/jobs/<job_id>')
@require_auth
def job_status(job_id):
    row = get_db().execute('SELECT * FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not row:
        return jsonify({'ok': False, 'motivo': 'job_no_encontrado'}), 404
    job = dict(row)
    if job.get('resumen'):
        job['resumen'] = json.loads(job['resumen'])
    return jsonify({'ok': True, **job})

@app.get('/api/cruces/<int:cruce_id>/download')
@require_auth
def cruce_download(cruce_id):
    db  = get_db()
    row = db.execute('SELECT excel_path FROM cruces WHERE id = ?', (cruce_id,)).fetchone()
    if not row or not row['excel_path']:
        return jsonify({'ok': False, 'motivo': 'sin_excel'}), 404
    path = Path(row['excel_path'])
    if not path.exists():
        return jsonify({'ok': False, 'motivo': 'archivo_no_encontrado'}), 404
    return send_file(str(path), as_attachment=True, download_name=f'cruce_{cruce_id}.xlsx')

@app.get('/api/cruces/<int:cruce_id>/export')
@require_auth
def cruce_export(cruce_id):
    from openpyxl.styles import Border, Side
    from openpyxl.utils  import get_column_letter

    db    = get_db()
    cruce = db.execute('SELECT * FROM cruces WHERE id = ?', (cruce_id,)).fetchone()
    if not cruce:
        return jsonify({'ok': False, 'motivo': 'no encontrado'}), 404

    lineas = db.execute(
        '''SELECT id, nro_pedido, punto_retiro, producto, gtin, sku, marca,
                  unidades, nodo_asignado, zona, stock_disponible,
                  estado, notas, alternativa_nodo, updated_at
           FROM lineas WHERE cruce_id = ?
           ORDER BY nodo_asignado, nro_pedido, producto''',
        (cruce_id,)
    ).fetchall()

    confirmadas = db.execute(
        'SELECT punto_retiro, receptor, entregado_at FROM entregas_confirmadas WHERE cruce_id = ?',
        (cruce_id,)
    ).fetchall()
    conf_map = {r['punto_retiro'].strip(): r for r in confirmadas}

    habilitadas = db.execute(
        'SELECT punto_retiro, nro_pedidos FROM entregas_habilitadas WHERE cruce_id = ?',
        (cruce_id,)
    ).fetchall()
    hab_map = {r['punto_retiro'].strip(): json.loads(r['nro_pedidos'] or '[]') for r in habilitadas}

    ESTADO_LABEL = {
        'PENDIENTE':     'Pendiente',
        'ENCONTRADO':    'Encontrado',
        'NO_ENCONTRADO': 'No encontrado',
        'MAL_STOCK':     'Mal stock',
        'REASIGNADO':    'Reasignado',
        'EN_REVISION':   'En revisión',
        'SIN_COBERTURA': 'Sin cobertura',
    }
    ESTADO_COLOR = {
        'ENCONTRADO':    '92D050',
        'NO_ENCONTRADO': 'FF4C4C',
        'MAL_STOCK':     'FFB84C',
        'EN_REVISION':   'B8B8FF',
        'SIN_COBERTURA': 'D3D3D3',
        'REASIGNADO':    'A9D18E',
        'PENDIENTE':     'FFFFFF',
    }

    HDR_FILL  = PatternFill(fill_type='solid', fgColor='2563EB')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
    HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN_SIDE = Side(style='thin', color='D0D5DD')
    THIN_BDR  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    DATA_FONT = Font(size=10)

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, HDR_ALIGN
            c.border = THIN_BDR
        ws.row_dimensions[1].height = 30

    def apply_row(ws, row_i, values, n_cols, fill=None):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.font   = DATA_FONT
            c.border = THIN_BDR
            c.alignment = Alignment(vertical='center', wrap_text=False)
            if fill:
                c.fill = fill

    def set_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    wb = openpyxl.Workbook()

    # ── Hoja 1: Detalle completo ─────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Detalle'

    h1 = ['N° Pedido', 'Punto de Retiro', 'Producto', 'GTIN', 'SKU', 'Marca',
          'Unidades', 'Sucursal Asignada', 'Zona', 'Stock Disp.', 'Estado',
          'Alternativa Propuesta', 'Notas', 'Última actualización']
    write_header(ws1, h1)

    for row_i, l in enumerate(lineas, 2):
        estado = l['estado'] or 'PENDIENTE'
        fill   = PatternFill(fill_type='solid', fgColor=ESTADO_COLOR.get(estado, 'FFFFFF'))
        apply_row(ws1, row_i, [
            l['nro_pedido'], l['punto_retiro'], l['producto'],
            l['gtin'], l['sku'], l['marca'], l['unidades'],
            l['nodo_asignado'], l['zona'], l['stock_disponible'],
            ESTADO_LABEL.get(estado, estado),
            l['alternativa_nodo'] or '',
            l['notas'] or '',
            (l['updated_at'] or '')[:16].replace('T', ' '),
        ], len(h1), fill=fill)

    set_widths(ws1, [14, 28, 36, 16, 14, 18, 9, 28, 14, 11, 14, 22, 24, 18])
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(h1))}1'

    # ── Hoja 2: Entregas por sucursal ────────────────────────────
    ws2 = wb.create_sheet('Entregas')

    h2 = ['Sucursal', 'N° Pedidos', 'Cant. Pedidos', 'Receptor', 'Hora de entrega', 'Estado entrega']
    write_header(ws2, h2)

    # Unión de habilitadas + confirmadas: ambas usan nodo_asignado como clave
    sucursales = sorted(set(hab_map.keys()) | set(conf_map.keys()))
    for row_i, suc in enumerate(sucursales, 2):
        conf    = conf_map.get(suc)
        pedidos = hab_map.get(suc, [])
        if not pedidos:
            pedidos_suc = sorted({l['nro_pedido'] for l in lineas
                                   if l['nodo_asignado'] and l['nodo_asignado'].strip() == suc and l['nro_pedido']})
        else:
            pedidos_suc = pedidos

        if conf:
            estado_ent = 'Entregado'
            fill       = PatternFill(fill_type='solid', fgColor='92D050')
            hora       = (conf['entregado_at'] or '')[:16].replace('T', ' ')
            receptor   = conf['receptor'] or ''
        else:
            estado_ent = 'Pendiente'
            fill       = PatternFill(fill_type='solid', fgColor='FFFFFF')
            hora       = ''
            receptor   = ''

        apply_row(ws2, row_i, [
            suc,
            ', '.join(str(p) for p in pedidos_suc),
            len(pedidos_suc),
            receptor,
            hora,
            estado_ent,
        ], len(h2), fill=fill)

    set_widths(ws2, [30, 40, 14, 24, 18, 16])
    ws2.freeze_panes = 'A2'

    # ── Hoja 3: Incidencias (NO_ENCONTRADO + MAL_STOCK) ──────────
    # Leemos desde eventos_estado para preservar la sucursal donde ocurrió cada
    # incidencia aunque la línea haya sido reasignada y resuelta después.
    ws3 = wb.create_sheet('Incidencias')

    h3 = ['N° Pedido', 'Punto de Retiro', 'Sucursal donde ocurrió', 'Producto',
          'GTIN', 'Unidades', 'Stock en sistema', 'Incidencia', 'Alternativa Propuesta', 'Notas']
    write_header(ws3, h3)

    eventos_inc = db.execute(
        '''SELECT e.estado_nuevo, e.nodo_asignado, e.created_at,
                  l.nro_pedido, l.punto_retiro, l.producto, l.gtin,
                  l.unidades, l.stock_disponible, l.alternativa_nodo, l.notas
           FROM eventos_estado e
           JOIN lineas l ON l.id = e.linea_id
           WHERE l.cruce_id = ? AND e.estado_nuevo IN ('NO_ENCONTRADO', 'MAL_STOCK')
           ORDER BY e.created_at''',
        (cruce_id,)
    ).fetchall()

    if eventos_inc:
        for row_i, ev in enumerate(eventos_inc, 2):
            estado = ev['estado_nuevo']
            fill   = PatternFill(fill_type='solid', fgColor=ESTADO_COLOR[estado])
            apply_row(ws3, row_i, [
                ev['nro_pedido'], ev['punto_retiro'], ev['nodo_asignado'],
                ev['producto'], ev['gtin'], ev['unidades'],
                ev['stock_disponible'],
                ESTADO_LABEL[estado],
                ev['alternativa_nodo'] or '',
                ev['notas'] or '',
            ], len(h3), fill=fill)
    else:
        ws3.cell(row=2, column=1, value='Sin incidencias registradas').font = Font(italic=True, color='888888')

    set_widths(ws3, [14, 28, 28, 36, 16, 9, 16, 16, 22, 24])
    ws3.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha    = (cruce['fecha'] or '')[:10]
    filename = f'reporte_cruce_{cruce_id}_{fecha}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Rutas lineas / estados cadete ───────────────────────────────────────────────

@app.get('/api/cruces/<int:cruce_id>/lineas')
@require_auth
def lineas_list(cruce_id):
    db = get_db()
    lineas = db.execute(
        'SELECT * FROM lineas WHERE cruce_id = ? ORDER BY nodo_asignado, nro_pedido',
        (cruce_id,)
    ).fetchall()
    result = []
    for l in lineas:
        d = dict(l)
        d['alternativas'] = json.loads(d['alternativas'] or '[]')
        result.append(d)
    return jsonify({'ok': True, 'lineas': result})

@app.patch('/api/cruces/<int:cruce_id>/lineas/<int:linea_id>')
@require_auth
def linea_update(cruce_id, linea_id):
    data = request.get_json(silent=True) or {}
    nuevo_estado = data.get('estado')
    motivo       = data.get('motivo', '')

    ESTADOS_VALIDOS = {'PENDIENTE', 'ENCONTRADO', 'NO_ENCONTRADO',
                       'MAL_STOCK', 'REASIGNADO', 'EN_REVISION', 'SIN_COBERTURA'}
    if nuevo_estado and nuevo_estado not in ESTADOS_VALIDOS:
        return jsonify({'ok': False, 'motivo': 'estado_invalido'}), 400

    db  = get_db()
    row = db.execute(
        'SELECT * FROM lineas WHERE id = ? AND cruce_id = ?', (linea_id, cruce_id)
    ).fetchone()
    if not row:
        return jsonify({'ok': False, 'motivo': 'no_encontrado'}), 404

    # Solo admin/operador pueden reasignar nodo
    if 'nodo_asignado' in data and g.rol not in ('admin', 'operador'):
        return jsonify({'ok': False, 'motivo': 'sin_permiso'}), 403

    updates = {}
    if nuevo_estado:
        updates['estado']     = nuevo_estado
        updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    if 'notas' in data:
        updates['notas'] = data['notas']
    if 'alternativa_nodo' in data:
        updates['alternativa_nodo'] = data['alternativa_nodo']
    if 'nodo_asignado' in data and g.rol in ('admin', 'operador'):
        updates['nodo_asignado'] = data['nodo_asignado'].strip() if data['nodo_asignado'] else None
        updates['updated_at']    = datetime.now(timezone.utc).isoformat()
    if 'stock_disponible' in data and g.rol in ('admin', 'operador'):
        val = data['stock_disponible']
        updates['stock_disponible'] = int(val) if val is not None else None

    if updates:
        set_clause = ', '.join(f'{k} = ?' for k in updates)
        db.execute(
            f'UPDATE lineas SET {set_clause} WHERE id = ?',
            list(updates.values()) + [linea_id]
        )

    if nuevo_estado and nuevo_estado != row['estado']:
        db.execute(
            '''INSERT INTO eventos_estado
               (linea_id, estado_anterior, estado_nuevo, motivo, origen, nodo_asignado)
               VALUES (?, ?, ?, ?, ?, ?)''',
            (linea_id, row['estado'], nuevo_estado, motivo, g.rol, row['nodo_asignado'])
        )

    db.commit()
    return jsonify({'ok': True})

@app.get('/api/cruces/<int:cruce_id>/progreso')
@require_auth
def cruce_progreso(cruce_id):
    db = get_db()
    row = db.execute(
        '''SELECT
             COUNT(*) as total,
             SUM(CASE WHEN estado = 'ENCONTRADO'    THEN 1 ELSE 0 END) as encontrado,
             SUM(CASE WHEN estado = 'NO_ENCONTRADO' THEN 1 ELSE 0 END) as no_encontrado,
             SUM(CASE WHEN estado = 'MAL_STOCK'     THEN 1 ELSE 0 END) as mal_stock,
             SUM(CASE WHEN estado = 'REASIGNADO'    THEN 1 ELSE 0 END) as reasignado,
             SUM(CASE WHEN estado = 'EN_REVISION'   THEN 1 ELSE 0 END) as en_revision,
             SUM(CASE WHEN estado = 'SIN_COBERTURA' THEN 1 ELSE 0 END) as sin_cobertura,
             SUM(CASE WHEN estado = 'PENDIENTE'     THEN 1 ELSE 0 END) as pendiente
           FROM lineas WHERE cruce_id = ?''',
        (cruce_id,)
    ).fetchone()
    d = dict(row)
    total = d['total'] or 1
    resueltos = (d['encontrado'] or 0) + (d['sin_cobertura'] or 0) + (d['no_encontrado'] or 0) + (d['reasignado'] or 0)
    d['pct_completado'] = round(resueltos / total * 100, 1)
    return jsonify({'ok': True, 'progreso': d})

# ── Entregas a puntos de retiro ─────────────────────────────────────────────────

@app.get('/api/cruces/<int:cruce_id>/entregas')
@require_auth
def cruces_entregas_list(cruce_id):
    db = get_db()

    # Sucursales fijas definidas en config.yaml
    try:
        with open(CONFIG_PATH) as f:
            _cfg = yaml.safe_load(f)
        sucursales_fijas = set(_cfg.get('sucursales_fijas', []))
    except Exception:
        sucursales_fijas = set()

    # Pedidos únicos por punto de retiro
    rows = db.execute(
        '''SELECT punto_retiro, nro_pedido
           FROM lineas
           WHERE cruce_id = ? AND punto_retiro IS NOT NULL AND punto_retiro != ''
           ORDER BY punto_retiro, nro_pedido''',
        (cruce_id,)
    ).fetchall()

    # Confirmaciones existentes
    confs = db.execute(
        'SELECT * FROM entregas_confirmadas WHERE cruce_id = ?', (cruce_id,)
    ).fetchall()
    conf_map = {c['punto_retiro']: dict(c) for c in confs}

    # Farmacias habilitadas con sus pedidos seleccionados
    hab_rows = db.execute(
        'SELECT punto_retiro, nro_pedidos FROM entregas_habilitadas WHERE cruce_id = ?',
        (cruce_id,)
    ).fetchall()
    hab_map     = {r['punto_retiro']: json.loads(r['nro_pedidos'] or '[]') for r in hab_rows}
    habilitadas = set(hab_map.keys())

    # Pedidos con al menos un producto ENCONTRADO por punto de retiro
    enc_rows = db.execute(
        '''SELECT DISTINCT punto_retiro, nro_pedido FROM lineas
           WHERE cruce_id = ? AND estado = 'ENCONTRADO'
             AND punto_retiro IS NOT NULL AND punto_retiro != ''
             AND nro_pedido   IS NOT NULL AND nro_pedido   != ''
        ''', (cruce_id,)
    ).fetchall()
    enc_map: dict = {}
    for r in enc_rows:
        enc_map.setdefault(r['punto_retiro'], set()).add(r['nro_pedido'])

    # Agrupar por punto_retiro (pedidos únicos)
    grupos: dict = {}
    for r in rows:
        pt = r['punto_retiro']
        if pt not in grupos:
            grupos[pt] = set()
        if r['nro_pedido']:
            grupos[pt].add(r['nro_pedido'])

    # Agregar todos los nodos del stock que no tienen punto_retiro propio
    nodo_rows = db.execute(
        '''SELECT DISTINCT nodo_asignado FROM lineas
           WHERE cruce_id = ? AND nodo_asignado IS NOT NULL AND nodo_asignado != ''
           ORDER BY nodo_asignado''',
        (cruce_id,)
    ).fetchall()
    for r in nodo_rows:
        if r['nodo_asignado'] not in grupos:
            grupos[r['nodo_asignado']] = set()

    # Agregar sucursales fijas del config aunque no estén en el stock
    for sf in sucursales_fijas:
        if sf not in grupos:
            grupos[sf] = set()

    # Agregar farmacias activadas o confirmadas manualmente (pedidos ingresados
    # a mano en el modal "Activar farmacia"): esos números nunca tocan `lineas`,
    # así que sin esto la entrega desaparece de la lista aunque esté confirmada
    # y firmada — ver CLAUDE.md, sección "tres conteos de pedidos".
    for pt in habilitadas:
        if pt not in grupos:
            grupos[pt] = set()
    for pt in conf_map:
        if pt not in grupos:
            grupos[pt] = set()

    # Mapeo punto_retiro → nodo_asignado más frecuente (para nombre_display)
    # Sólo se usa el nombre del nodo si comparte el número de calle con el
    # punto_retiro: si el pedido se fulfillea con stock de OTRA sucursal
    # (algo normal en este sistema), el nodo más frecuente puede ser una
    # sucursal completamente distinta — mostrar su nombre pisaría la
    # identidad real del punto de retiro y haría ver "duplicada" una
    # entrega que en realidad es otro destino.
    nodo_nombre_rows = db.execute(
        '''SELECT punto_retiro, nodo_asignado, COUNT(*) as cnt
           FROM lineas
           WHERE cruce_id = ?
             AND punto_retiro  != '' AND punto_retiro  IS NOT NULL
             AND nodo_asignado != '' AND nodo_asignado IS NOT NULL
           GROUP BY punto_retiro, nodo_asignado
           ORDER BY punto_retiro, cnt DESC''',
        (cruce_id,)
    ).fetchall()

    def _mismo_lugar(punto_retiro, nodo_asignado):
        nums_pt   = set(re.findall(r'\d{2,}', punto_retiro))
        nums_nodo = set(re.findall(r'\d{2,}', nodo_asignado))
        return bool(nums_pt & nums_nodo)

    nombre_display_map = {}
    for r in nodo_nombre_rows:
        pt = r['punto_retiro']
        if pt not in nombre_display_map:
            nombre_display_map[pt] = r['nodo_asignado'] if _mismo_lugar(pt, r['nodo_asignado']) else pt

    # Teléfonos — match exacto por nombre_display (nombre de nodo)
    tel_rows  = db.execute('SELECT nombre, telefono FROM nodos_config WHERE telefono != ""').fetchall()
    nodos_map = {r['nombre']: r['telefono'] for r in tel_rows}

    result = []
    for pt, pedidos in sorted(grupos.items()):
        conf            = conf_map.get(pt)
        nombre_display  = nombre_display_map.get(pt, pt)
        todos_sorted    = sorted(list(pedidos))
        pedidos_activos = hab_map.get(pt, [])
        pedidos_enc     = sorted(list(enc_map.get(pt, set())))
        result.append({
            'punto_retiro':        pt,
            'nombre_display':      nombre_display,
            'nro_pedidos':         todos_sorted,
            'total_pedidos':       len(pedidos),
            'tiene_pedidos':       len(pedidos) > 0,
            'activa':              pt in habilitadas,
            'pedidos_activos':     pedidos_activos,
            'pedidos_encontrados': pedidos_enc,
            'confirmada':          conf is not None,
            'receptor':            conf['receptor']     if conf else None,
            'entregado_at':        conf['entregado_at'] if conf else None,
            'firma_base64':        conf['firma_base64'] if conf else None,
            'entrega_id':          conf['id']           if conf else None,
            'telefono':            _match_telefono(nombre_display, nodos_map),
        })

    return jsonify({'ok': True, 'entregas': result, 'total': len(result)})


@app.post('/api/cruces/<int:cruce_id>/entregas/activar')
@require_auth
def cruces_entregas_activar(cruce_id):
    data         = request.get_json(silent=True) or {}
    punto_retiro = data.get('punto_retiro', '').strip()
    activa       = data.get('activa', True)
    nro_pedidos  = data.get('nro_pedidos', [])   # lista de pedidos seleccionados

    if not punto_retiro:
        return jsonify({'ok': False, 'motivo': 'falta_punto_retiro'}), 400

    db = get_db()
    if activa:
        db.execute(
            '''INSERT INTO entregas_habilitadas (cruce_id, punto_retiro, nro_pedidos)
               VALUES (?, ?, ?)
               ON CONFLICT(cruce_id, punto_retiro)
               DO UPDATE SET nro_pedidos = excluded.nro_pedidos''',
            (cruce_id, punto_retiro, json.dumps(nro_pedidos))
        )
    else:
        db.execute(
            'DELETE FROM entregas_habilitadas WHERE cruce_id = ? AND punto_retiro = ?',
            (cruce_id, punto_retiro)
        )
    db.commit()
    return jsonify({'ok': True, 'activa': activa})


@app.post('/api/cruces/<int:cruce_id>/entregas')
@require_auth
def cruces_entregas_confirmar(cruce_id):
    data         = request.get_json(silent=True) or {}
    punto_retiro = data.get('punto_retiro', '').strip()
    receptor     = data.get('receptor', '').strip()
    firma_b64    = data.get('firma_base64', '')

    if not punto_retiro or not receptor:
        return jsonify({'ok': False, 'motivo': 'campos_requeridos'}), 400

    db  = get_db()
    now = datetime.now(timezone.utc).isoformat()

    existing = db.execute(
        'SELECT id FROM entregas_confirmadas WHERE cruce_id = ? AND punto_retiro = ?',
        (cruce_id, punto_retiro)
    ).fetchone()

    if existing:
        db.execute(
            '''UPDATE entregas_confirmadas
               SET receptor = ?, firma_base64 = ?, entregado_at = ?, usuario_id = ?
               WHERE id = ?''',
            (receptor, firma_b64, now, g.user_id, existing['id'])
        )
    else:
        db.execute(
            '''INSERT INTO entregas_confirmadas
               (cruce_id, punto_retiro, receptor, firma_base64, entregado_at, usuario_id)
               VALUES (?, ?, ?, ?, ?, ?)''',
            (cruce_id, punto_retiro, receptor, firma_b64, now, g.user_id)
        )

    db.commit()
    return jsonify({'ok': True})

# ── Nodos config (teléfonos de farmacias) ───────────────────────────────────────

@app.get('/api/nodos-config')
@require_auth
def nodos_config_list():
    db = get_db()
    # Teléfonos guardados
    rows = db.execute('SELECT nombre, telefono FROM nodos_config ORDER BY nombre').fetchall()
    # Nombres reales de nodos (nodo_asignado) — son los que aparecen en las tiles
    conocidos = db.execute(
        '''SELECT DISTINCT nodo_asignado AS nombre FROM lineas
           WHERE nodo_asignado IS NOT NULL AND nodo_asignado != ''
           ORDER BY nombre'''
    ).fetchall()
    return jsonify({
        'ok':     True,
        'nodos':  [dict(r) for r in rows],
        'conocidos': [r['nombre'] for r in conocidos],
    })

@app.put('/api/nodos-config')
@require_admin
def nodos_config_upsert():
    data     = request.get_json(silent=True) or {}
    nombre   = (data.get('nombre') or '').strip()
    telefono = (data.get('telefono') or '').strip()
    if not nombre:
        return jsonify({'ok': False, 'motivo': 'nombre_requerido'}), 400
    db = get_db()
    db.execute(
        '''INSERT INTO nodos_config (nombre, telefono) VALUES (?, ?)
           ON CONFLICT(nombre) DO UPDATE SET telefono = excluded.telefono''',
        (nombre, telefono)
    )
    db.commit()
    return jsonify({'ok': True})

# ── Puntos de retiro: links públicos para que la farmacia vea sus pedidos ────────
# El token identifica un punto_retiro exacto (nunca nombre_display — ver CLAUDE.md),
# así que dos entregas que muestran el mismo nombre por coincidencia no pueden
# pisarse el link.

@app.get('/api/config/puntos-retiro')
@require_admin
def puntos_retiro_list():
    db = get_db()
    try:
        with open(CONFIG_PATH) as f:
            _cfg = yaml.safe_load(f)
        sucursales_fijas = set(_cfg.get('sucursales_fijas', []))
    except Exception:
        sucursales_fijas = set()

    conocidos = db.execute(
        '''SELECT DISTINCT punto_retiro FROM lineas
           WHERE punto_retiro IS NOT NULL AND punto_retiro != ''
           ORDER BY punto_retiro'''
    ).fetchall()
    puntos = sorted(sucursales_fijas | {r['punto_retiro'] for r in conocidos})

    tok_rows = db.execute('SELECT punto_retiro, token FROM puntos_retiro_tokens').fetchall()
    tok_map  = {r['punto_retiro']: r['token'] for r in tok_rows}

    return jsonify({
        'ok': True,
        'puntos': [{'punto_retiro': p, 'token': tok_map.get(p)} for p in puntos],
    })

@app.post('/api/config/puntos-retiro/token')
@require_admin
def puntos_retiro_generar_token():
    data         = request.get_json(silent=True) or {}
    punto_retiro = (data.get('punto_retiro') or '').strip()
    if not punto_retiro:
        return jsonify({'ok': False, 'motivo': 'punto_retiro_requerido'}), 400

    token = secrets.token_urlsafe(16)
    db = get_db()
    db.execute(
        '''INSERT INTO puntos_retiro_tokens (punto_retiro, token) VALUES (?, ?)
           ON CONFLICT(punto_retiro)
           DO UPDATE SET token = excluded.token, created_at = CURRENT_TIMESTAMP''',
        (punto_retiro, token)
    )
    db.commit()
    return jsonify({'ok': True, 'token': token})

RETIRO_VENTANA_DIAS = 15

def _mask_dni(dni: str) -> str:
    """Oculta todo salvo los últimos 3 dígitos: '30123456' → '*****456'."""
    dni = (dni or '').strip()
    if len(dni) <= 3:
        return dni
    return '*' * (len(dni) - 3) + dni[-3:]

def _resolver_token_retiro(token):
    db  = get_db()
    row = db.execute(
        'SELECT punto_retiro FROM puntos_retiro_tokens WHERE token = ?', (token,)
    ).fetchone()
    return row['punto_retiro'] if row else None

@app.get('/api/retiro/<token>')
def retiro_publico(token):
    """Vista pública (sin auth): qué pedidos de SU punto_retiro exacto ya bajó
    y firmó el cadete (están en la góndola) en los últimos RETIRO_VENTANA_DIAS,
    más los del cruce más reciente si todavía no se confirmó esa entrega."""
    db = get_db()
    punto_retiro = _resolver_token_retiro(token)
    if not punto_retiro:
        return jsonify({'ok': False, 'motivo': 'token_invalido'}), 404

    def _pedidos_del_cruce(cruce_id):
        """Pedidos activos (con fallback a los del cruce — regla dura del
        proyecto, ver CLAUDE.md) de este punto_retiro, con sus productos y cliente."""
        hab_row = db.execute(
            'SELECT nro_pedidos FROM entregas_habilitadas WHERE cruce_id = ? AND punto_retiro = ?',
            (cruce_id, punto_retiro)
        ).fetchone()
        activos = json.loads(hab_row['nro_pedidos']) if hab_row and hab_row['nro_pedidos'] else []

        lineas = db.execute(
            '''SELECT nro_pedido, producto, marca, unidades,
                      cliente_nombre, cliente_dni, cliente_telefono
               FROM lineas
               WHERE cruce_id = ? AND punto_retiro = ?
                 AND nro_pedido IS NOT NULL AND nro_pedido != ''
               ORDER BY nro_pedido''',
            (cruce_id, punto_retiro)
        ).fetchall()
        por_pedido: dict = {}
        cliente_por_pedido: dict = {}
        for l in lineas:
            por_pedido.setdefault(l['nro_pedido'], []).append({
                'producto': l['producto'], 'marca': l['marca'], 'unidades': l['unidades'],
            })
            if l['nro_pedido'] not in cliente_por_pedido and l['cliente_nombre']:
                cliente_por_pedido[l['nro_pedido']] = {
                    'nombre':   l['cliente_nombre'],
                    'dni':      _mask_dni(l['cliente_dni']),
                    'telefono': l['cliente_telefono'],
                }
        mostrar = activos if activos else sorted(por_pedido.keys())
        return {
            np: (por_pedido.get(np, []), cliente_por_pedido.get(np))
            for np in mostrar
        }

    cutoff = (datetime.now(timezone.utc) - timedelta(days=RETIRO_VENTANA_DIAS)).isoformat()
    confs = db.execute(
        '''SELECT cruce_id, receptor, entregado_at FROM entregas_confirmadas
           WHERE punto_retiro = ? AND entregado_at >= ?
           ORDER BY entregado_at DESC''',
        (punto_retiro, cutoff)
    ).fetchall()

    retirados = {
        r['nro_pedido'] for r in db.execute(
            'SELECT nro_pedido FROM pedidos_retirados WHERE punto_retiro = ?', (punto_retiro,)
        ).fetchall()
    }

    pedidos: dict = {}
    for c in confs:
        for np, (productos, cliente) in _pedidos_del_cruce(c['cruce_id']).items():
            if np in pedidos or np in retirados:
                continue  # ya está de una confirmación más reciente, o ya lo retiraron
            pedidos[np] = {
                'nro_pedido':    np,
                'confirmado':    True,
                'confirmado_en': c['entregado_at'],
                'receptor':      c['receptor'],
                'productos':     productos,
                'cliente':       cliente,
            }

    # Pendientes: sólo el cruce más reciente, y sólo si todavía no se confirmó
    # esa entrega para este punto_retiro (para no arrastrar cruces viejos sin firmar).
    ultimo_cruce = db.execute('SELECT id FROM cruces ORDER BY id DESC LIMIT 1').fetchone()
    if ultimo_cruce and not any(c['cruce_id'] == ultimo_cruce['id'] for c in confs):
        for np, (productos, cliente) in _pedidos_del_cruce(ultimo_cruce['id']).items():
            if np not in pedidos and np not in retirados:
                pedidos[np] = {
                    'nro_pedido':    np,
                    'confirmado':    False,
                    'confirmado_en': None,
                    'receptor':      None,
                    'productos':     productos,
                    'cliente':       cliente,
                }

    resultado = sorted(pedidos.values(), key=lambda p: (not p['confirmado'], p['nro_pedido']))

    return jsonify({'ok': True, 'punto_retiro': punto_retiro, 'pedidos': resultado})

@app.post('/api/admin/backfill-cliente/<int:cruce_id>')
@require_admin
def _backfill_cliente_TEMPORAL(cruce_id):
    """Migración puntual: completa cliente_nombre/dni/telefono de un cruce ya
    existente releyendo su archivo de pedidos original. Se borra después de usarla."""
    from src.loader  import cargar_archivo
    from src.matcher import mapear_columnas_pedidos

    db  = get_db()
    cru = db.execute('SELECT archivo_pedidos, excel_path FROM cruces WHERE id = ?', (cruce_id,)).fetchone()
    if not cru:
        return jsonify({'ok': False, 'motivo': 'cruce_no_encontrado'}), 404

    candidatos = []
    if cru['excel_path']:
        misma_carpeta = Path(cru['excel_path']).parent / cru['archivo_pedidos']
        if misma_carpeta.exists():
            candidatos.append(misma_carpeta)
    if not candidatos:
        candidatos = list(UPLOADS_DIR.rglob(cru['archivo_pedidos']))
    if not candidatos:
        return jsonify({'ok': False, 'motivo': 'archivo_no_encontrado', 'buscado': cru['archivo_pedidos']}), 404

    with open(str(CONFIG_PATH), 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    df = cargar_archivo(str(candidatos[0]))
    mapa = mapear_columnas_pedidos(df, config)
    col_nro = mapa.get('nro_pedido')
    col_nom = mapa.get('cliente_nombre')
    col_dni = mapa.get('cliente_dni')
    col_tel = mapa.get('cliente_telefono')
    if not col_nro or not (col_nom or col_dni or col_tel):
        return jsonify({'ok': False, 'motivo': 'sin_columnas_cliente_o_pedido'}), 400

    actualizados = 0
    vistos = set()
    for _, fila in df.iterrows():
        nro = str(fila.get(col_nro, '') or '').strip()
        if not nro or nro in vistos:
            continue
        vistos.add(nro)
        nombre   = str(fila.get(col_nom, '') or '').strip() if col_nom else ''
        dni      = str(fila.get(col_dni, '') or '').strip() if col_dni else ''
        telefono = str(fila.get(col_tel, '') or '').strip() if col_tel else ''
        if not (nombre or dni or telefono):
            continue
        cur = db.execute(
            '''UPDATE lineas SET cliente_nombre = ?, cliente_dni = ?, cliente_telefono = ?
               WHERE cruce_id = ? AND nro_pedido = ? AND (cliente_nombre IS NULL OR cliente_nombre = '')''',
            (nombre, dni, telefono, cruce_id, nro)
        )
        actualizados += cur.rowcount
    db.commit()
    return jsonify({'ok': True, 'archivo': str(candidatos[0]), 'lineas_actualizadas': actualizados})

@app.post('/api/retiro/<token>/marcar-retirado')
def retiro_marcar_retirado(token):
    """Público (sin auth): la farmacia marca que el cliente ya se llevó el pedido."""
    punto_retiro = _resolver_token_retiro(token)
    if not punto_retiro:
        return jsonify({'ok': False, 'motivo': 'token_invalido'}), 404

    data       = request.get_json(silent=True) or {}
    nro_pedido = (data.get('nro_pedido') or '').strip()
    if not nro_pedido:
        return jsonify({'ok': False, 'motivo': 'nro_pedido_requerido'}), 400

    db = get_db()
    db.execute(
        '''INSERT INTO pedidos_retirados (punto_retiro, nro_pedido) VALUES (?, ?)
           ON CONFLICT(punto_retiro, nro_pedido) DO NOTHING''',
        (punto_retiro, nro_pedido)
    )
    db.commit()
    return jsonify({'ok': True})

# ── Incidencias (MAL_STOCK / NO_ENCONTRADO / EN_REVISION / REASIGNADO) ──────────

@app.get('/api/cruces/<int:cruce_id>/incidencias')
@require_auth
def cruce_incidencias(cruce_id):
    db = get_db()
    lineas = db.execute(
        '''SELECT * FROM lineas
           WHERE cruce_id = ? AND estado IN ('MAL_STOCK','NO_ENCONTRADO','EN_REVISION','REASIGNADO')
           ORDER BY updated_at DESC''',
        (cruce_id,)
    ).fetchall()
    result = []
    for l in lineas:
        d = dict(l)
        d['alternativas'] = json.loads(d['alternativas'] or '[]')
        result.append(d)
    return jsonify({'ok': True, 'incidencias': result, 'total': len(result)})

@app.get('/api/cruces/<int:cruce_id>/stream')
@require_auth
def cruce_stream(cruce_id):
    """SSE: emite eventos de incidencia nuevas cada 8s. Reemplaza el poll de 15s."""
    import sqlite3 as _sqlite3

    def generate():
        conn = _sqlite3.connect(str(DB_PATH))
        conn.row_factory = _sqlite3.Row
        desde = datetime.now(timezone.utc).isoformat()
        try:
            while True:
                rows = conn.execute(
                    '''SELECT id, nro_pedido, producto, nodo_asignado, estado, updated_at, punto_retiro
                       FROM lineas WHERE cruce_id = ? AND estado IN ('MAL_STOCK','NO_ENCONTRADO')
                       AND updated_at > ? ORDER BY updated_at ASC LIMIT 20''',
                    (cruce_id, desde)
                ).fetchall()
                if rows:
                    desde = datetime.now(timezone.utc).isoformat()
                    for r in rows:
                        yield f'data: {json.dumps(dict(r), ensure_ascii=False)}\n\n'
                else:
                    yield ': ping\n\n'
                time.sleep(8)
        except GeneratorExit:
            pass
        finally:
            conn.close()

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )

@app.get('/api/cruces/<int:cruce_id>/notificaciones')
@require_auth
def cruce_notificaciones(cruce_id):
    """Devuelve MAL_STOCK / NO_ENCONTRADO nuevos desde ?desde=<iso-timestamp>."""
    desde  = request.args.get('desde', '')
    db     = get_db()
    query  = '''SELECT id, nro_pedido, producto, nodo_asignado, estado, updated_at, punto_retiro
                FROM lineas
                WHERE cruce_id = ? AND estado IN ('MAL_STOCK','NO_ENCONTRADO')'''
    params = [cruce_id]
    if desde:
        query  += ' AND updated_at > ?'
        params.append(desde)
    query += ' ORDER BY updated_at DESC LIMIT 20'
    rows   = db.execute(query, params).fetchall()
    return jsonify({'ok': True, 'nuevas': [dict(r) for r in rows]})

# ── Analytics ───────────────────────────────────────────────────────────────────

def _analytics_query(db, desde=None, hasta=None):
    """Ejecuta todas las queries de analytics con filtro de rango de fechas."""
    params_c  = []
    params_l  = []
    where_c   = ''
    where_l   = ''

    if desde:
        where_c += ' AND c.created_at >= ?'
        where_l += ' AND cr.created_at >= ?'
        params_c.append(desde)
        params_l.append(desde)
    if hasta:
        # incluir todo el día "hasta"
        hasta_fin = hasta + 'T23:59:59'
        where_c  += ' AND c.created_at <= ?'
        where_l  += ' AND cr.created_at <= ?'
        params_c.append(hasta_fin)
        params_l.append(hasta_fin)

    totales = db.execute(
        f'''SELECT COUNT(*) as total_cruces,
                   COALESCE(AVG(c.pct_cobertura), 0) as cobertura_promedio,
                   COALESCE(SUM(c.total_lineas), 0)  as total_lineas,
                   COALESCE(SUM(c.con_cobertura), 0)  as total_con_cobertura,
                   COALESCE(SUM(c.sin_cobertura), 0)  as total_sin_cobertura,
                   COUNT(DISTINCT l.nro_pedido) as total_pedidos,
                   COALESCE(SUM(CASE WHEN l.estado IN ('MAL_STOCK','NO_ENCONTRADO') THEN 1 ELSE 0 END), 0) as total_incidencias
            FROM cruces c
            LEFT JOIN lineas l ON l.cruce_id = c.id
            WHERE 1=1 {where_c}''',
        params_c
    ).fetchone()

    evolucion = db.execute(
        f'''SELECT DATE(c.created_at) as dia,
                   ROUND(AVG(c.pct_cobertura), 1) as cobertura,
                   SUM(c.total_lineas) as lineas,
                   COUNT(*) as cruces
            FROM cruces c WHERE 1=1 {where_c}
            GROUP BY dia ORDER BY dia DESC LIMIT 30''',
        params_c
    ).fetchall()

    top_productos = db.execute(
        f'''SELECT l.producto, COUNT(*) as veces,
                   SUM(l.unidades) as unidades
            FROM lineas l
            JOIN cruces cr ON cr.id = l.cruce_id
            WHERE l.producto IS NOT NULL AND l.producto != '' {where_l}
            GROUP BY l.producto ORDER BY veces DESC LIMIT 10''',
        params_l
    ).fetchall()

    top_sin_stock = db.execute(
        f'''SELECT l.producto, COUNT(*) as veces,
                   SUM(l.unidades) as unidades
            FROM lineas l
            JOIN cruces cr ON cr.id = l.cruce_id
            WHERE l.nodo_asignado IS NULL AND l.producto IS NOT NULL {where_l}
            GROUP BY l.producto ORDER BY veces DESC LIMIT 10''',
        params_l
    ).fetchall()

    top_farmacias = db.execute(
        f'''SELECT l.nodo_asignado as farmacia, COUNT(*) as lineas,
                   COUNT(DISTINCT l.nro_pedido) as pedidos
            FROM lineas l
            JOIN cruces cr ON cr.id = l.cruce_id
            WHERE l.nodo_asignado IS NOT NULL AND l.nodo_asignado != ''
              AND l.nodo_asignado != '— SIN COBERTURA —' {where_l}
            GROUP BY l.nodo_asignado ORDER BY lineas DESC LIMIT 10''',
        params_l
    ).fetchall()

    farmacias_problema = db.execute(
        f'''SELECT l.nodo_asignado as farmacia,
                   SUM(CASE WHEN l.estado IN ('MAL_STOCK','NO_ENCONTRADO') THEN 1 ELSE 0 END) as incidencias,
                   COUNT(*) as total_asignadas
            FROM lineas l
            JOIN cruces cr ON cr.id = l.cruce_id
            WHERE l.nodo_asignado IS NOT NULL AND l.nodo_asignado != ''
              AND l.nodo_asignado != '— SIN COBERTURA —' {where_l}
            GROUP BY l.nodo_asignado
            HAVING incidencias > 0
            ORDER BY incidencias DESC LIMIT 10''',
        params_l
    ).fetchall()

    top_puntos_retiro = db.execute(
        f'''SELECT l.punto_retiro, COUNT(DISTINCT l.nro_pedido) as pedidos
            FROM lineas l
            JOIN cruces cr ON cr.id = l.cruce_id
            WHERE l.punto_retiro IS NOT NULL AND l.punto_retiro != '' {where_l}
            GROUP BY l.punto_retiro ORDER BY pedidos DESC LIMIT 10''',
        params_l
    ).fetchall()

    return {
        'totales':           dict(totales),
        'evolucion':         [dict(r) for r in evolucion],
        'top_productos':     [dict(r) for r in top_productos],
        'top_sin_stock':     [dict(r) for r in top_sin_stock],
        'top_farmacias':     [dict(r) for r in top_farmacias],
        'farmacias_problema':[dict(r) for r in farmacias_problema],
        'top_puntos_retiro': [dict(r) for r in top_puntos_retiro],
    }


@app.get('/api/analytics')
@require_auth
def analytics():
    db    = get_db()
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    data  = _analytics_query(db, desde or None, hasta or None)
    return jsonify({'ok': True, **data})


@app.get('/api/analytics/export')
@require_auth
def analytics_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    db    = get_db()
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    data  = _analytics_query(db, desde or None, hasta or None)

    wb = openpyxl.Workbook()

    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1D4ED8')
    thin      = Side(style='thin', color='D1D5DB')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(row)
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len    = max(
                (len(str(ws.cell(row=r, column=col_idx).value or ''))
                 for r in range(1, ws.max_row + 1)),
                default=8
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        return ws

    del wb['Sheet']  # quitar hoja default

    # Resumen
    t   = data['totales']
    ws0 = wb.create_sheet('Resumen')
    periodo_str = f"{desde or 'inicio'} → {hasta or 'hoy'}"
    ws0['A1'] = 'Período'
    ws0['B1'] = periodo_str
    ws0['A1'].font = Font(bold=True)
    ws0.append(['Cruces realizados',        t['total_cruces']])
    ws0.append(['Cobertura promedio (%)',    round(t['cobertura_promedio'], 1)])
    ws0.append(['Líneas procesadas',         t['total_lineas']])
    ws0.append(['Pedidos únicos',            t['total_pedidos']])
    ws0.append(['Líneas con cobertura',      t['total_con_cobertura']])
    ws0.append(['Líneas sin cobertura',      t['total_sin_cobertura']])
    ws0.column_dimensions['A'].width = 28
    ws0.column_dimensions['B'].width = 20

    make_sheet(
        'Evolución',
        ['Día', 'Cobertura (%)', 'Líneas', 'Cruces'],
        [[r['dia'], r['cobertura'], r['lineas'], r['cruces']]
         for r in reversed(data['evolucion'])]
    )
    make_sheet(
        'Top Productos',
        ['Producto', 'Veces pedido', 'Unidades totales'],
        [[r['producto'], r['veces'], r['unidades']] for r in data['top_productos']]
    )
    make_sheet(
        'Sin Stock',
        ['Producto sin cobertura', 'Veces', 'Unidades afectadas'],
        [[r['producto'], r['veces'], r['unidades']] for r in data['top_sin_stock']]
    )
    make_sheet(
        'Top Farmacias',
        ['Farmacia', 'Líneas asignadas', 'Pedidos'],
        [[r['farmacia'], r['lineas'], r['pedidos']] for r in data['top_farmacias']]
    )
    make_sheet(
        'Farmacias Problema',
        ['Farmacia', 'Incidencias', 'Total asignadas'],
        [[r['farmacia'], r['incidencias'], r['total_asignadas']]
         for r in data['farmacias_problema']]
    )
    make_sheet(
        'Puntos de Retiro',
        ['Punto de retiro', 'Pedidos'],
        [[r['punto_retiro'], r['pedidos']] for r in data['top_puntos_retiro']]
    )

    # ── Filtro de rango reutilizado para Incidencias y Entregas ──
    where_c, params_c = '', []
    where_cr, params_cr = '', []  # mismo filtro, alias "cr" (cruces vía join)
    if desde:
        where_c   += ' AND c.created_at >= ?'
        where_cr  += ' AND cr.created_at >= ?'
        params_c.append(desde)
        params_cr.append(desde)
    if hasta:
        hasta_fin = hasta + 'T23:59:59'
        where_c   += ' AND c.created_at <= ?'
        where_cr  += ' AND cr.created_at <= ?'
        params_c.append(hasta_fin)
        params_cr.append(hasta_fin)

    ESTADO_LABEL = {
        'NO_ENCONTRADO': 'No encontrado',
        'MAL_STOCK':     'Mal stock',
    }

    # ── Hoja: Incidencias (NO_ENCONTRADO + MAL_STOCK en todo el rango) ──
    eventos_inc = db.execute(
        f'''SELECT DATE(cr.created_at) as fecha, cr.id as cruce_id,
                   e.estado_nuevo, e.nodo_asignado,
                   l.nro_pedido, l.punto_retiro, l.producto, l.gtin,
                   l.unidades, l.stock_disponible, l.alternativa_nodo, l.notas
            FROM eventos_estado e
            JOIN lineas l ON l.id = e.linea_id
            JOIN cruces cr ON cr.id = l.cruce_id
            WHERE e.estado_nuevo IN ('NO_ENCONTRADO', 'MAL_STOCK') {where_cr}
            ORDER BY cr.created_at DESC, e.created_at''',
        params_cr
    ).fetchall()

    make_sheet(
        'Incidencias',
        ['Fecha', 'Cruce', 'N° Pedido', 'Punto de Retiro', 'Sucursal donde ocurrió',
         'Producto', 'GTIN', 'Unidades', 'Stock en sistema', 'Incidencia',
         'Alternativa Propuesta', 'Notas'],
        [[ev['fecha'], ev['cruce_id'], ev['nro_pedido'], ev['punto_retiro'],
          ev['nodo_asignado'], ev['producto'], ev['gtin'], ev['unidades'],
          ev['stock_disponible'], ESTADO_LABEL.get(ev['estado_nuevo'], ev['estado_nuevo']),
          ev['alternativa_nodo'] or '', ev['notas'] or '']
         for ev in eventos_inc]
    )

    # ── Hoja: Entregas por sucursal (todos los cruces del rango) ──
    # Misma lógica que /api/cruces/<id>/export: punto_retiro es la identidad de
    # la entrega; pedidos_activos (entregas_habilitadas) es la fuente de verdad,
    # con fallback a nodo_asignado en lineas solo si no hubo activación manual.
    cruces_rango = db.execute(
        f'SELECT id, created_at FROM cruces c WHERE 1=1 {where_c} ORDER BY c.created_at DESC',
        params_c
    ).fetchall()

    entregas_rows = []
    for cr in cruces_rango:
        cid      = cr['id']
        fecha_cr = (cr['created_at'] or '')[:10]

        confirmadas = db.execute(
            'SELECT punto_retiro, receptor, entregado_at FROM entregas_confirmadas WHERE cruce_id = ?',
            (cid,)
        ).fetchall()
        conf_map = {r['punto_retiro'].strip(): r for r in confirmadas}

        habilitadas = db.execute(
            'SELECT punto_retiro, nro_pedidos FROM entregas_habilitadas WHERE cruce_id = ?',
            (cid,)
        ).fetchall()
        hab_map = {r['punto_retiro'].strip(): json.loads(r['nro_pedidos'] or '[]') for r in habilitadas}

        if not hab_map and not conf_map:
            continue

        lineas_cr = db.execute(
            'SELECT nro_pedido, nodo_asignado FROM lineas WHERE cruce_id = ?', (cid,)
        ).fetchall()

        for suc in sorted(set(hab_map.keys()) | set(conf_map.keys())):
            conf    = conf_map.get(suc)
            pedidos = hab_map.get(suc, [])
            if not pedidos:
                pedidos_suc = sorted({l['nro_pedido'] for l in lineas_cr
                                       if l['nodo_asignado'] and l['nodo_asignado'].strip() == suc and l['nro_pedido']})
            else:
                pedidos_suc = pedidos

            if conf:
                estado_ent = 'Entregado'
                hora       = (conf['entregado_at'] or '')[:16].replace('T', ' ')
                receptor   = conf['receptor'] or ''
            else:
                estado_ent = 'Pendiente'
                hora       = ''
                receptor   = ''

            entregas_rows.append([
                fecha_cr, cid, suc,
                ', '.join(str(p) for p in pedidos_suc),
                len(pedidos_suc), receptor, hora, estado_ent,
            ])

    make_sheet(
        'Entregas',
        ['Fecha', 'Cruce', 'Sucursal', 'N° Pedidos', 'Cant. Pedidos',
         'Receptor', 'Hora de entrega', 'Estado entrega'],
        entregas_rows
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha_str = datetime.now().strftime('%Y%m%d')
    nombre    = f'analytics_{fecha_str}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=nombre,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

# ── Incidencias operativas ──────────────────────────────────────────────────────

TIPOS_INCIDENCIA = {'pedido_perdido', 'mal_armado', 'demora', 'reclamo', 'otro'}

@app.get('/api/incidencias')
@require_auth
def incidencias_list():
    db    = get_db()
    tipo  = request.args.get('tipo', '')
    estado = request.args.get('estado', '')
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')

    conds  = ['1=1']
    params = []
    if tipo:
        conds.append('tipo = ?'); params.append(tipo)
    if estado:
        conds.append('estado = ?'); params.append(estado)
    if desde:
        conds.append('created_at >= ?'); params.append(desde)
    if hasta:
        conds.append('created_at <= ?'); params.append(hasta + 'T23:59:59')

    rows = db.execute(
        f'SELECT * FROM incidencias_operativas WHERE {" AND ".join(conds)} ORDER BY created_at DESC',
        params
    ).fetchall()
    return jsonify({'ok': True, 'incidencias': [dict(r) for r in rows]})


@app.post('/api/incidencias')
@require_auth
def incidencia_create():
    data = request.get_json(silent=True) or {}
    tipo        = data.get('tipo', 'otro')
    descripcion = (data.get('descripcion') or '').strip()
    if not descripcion:
        return jsonify({'ok': False, 'motivo': 'descripcion_requerida'}), 400
    if tipo not in TIPOS_INCIDENCIA:
        return jsonify({'ok': False, 'motivo': 'tipo_invalido'}), 400

    db  = get_db()
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        '''INSERT INTO incidencias_operativas
           (tipo, nro_pedido, sucursal, descripcion, estado, creado_por, created_at, updated_at)
           VALUES (?, ?, ?, ?, 'abierta', ?, ?, ?)''',
        (tipo, data.get('nro_pedido', ''), data.get('sucursal', ''),
         descripcion, g.usuario, now, now)
    )
    db.commit()
    return jsonify({'ok': True, 'id': cur.lastrowid}), 201


@app.patch('/api/incidencias/<int:inc_id>')
@require_auth
def incidencia_update(inc_id):
    data = request.get_json(silent=True) or {}
    db   = get_db()
    row  = db.execute('SELECT * FROM incidencias_operativas WHERE id = ?', (inc_id,)).fetchone()
    if not row:
        return jsonify({'ok': False, 'motivo': 'no_encontrado'}), 404

    updates = {}
    if 'estado' in data and data['estado'] in ('abierta', 'resuelta'):
        updates['estado'] = data['estado']
    if 'resolucion' in data:
        updates['resolucion'] = data['resolucion']
    if 'descripcion' in data and data['descripcion'].strip():
        updates['descripcion'] = data['descripcion'].strip()
    if 'sucursal' in data:
        updates['sucursal'] = data['sucursal']
    if 'nro_pedido' in data:
        updates['nro_pedido'] = data['nro_pedido']

    if updates:
        updates['updated_at'] = datetime.now(timezone.utc).isoformat()
        set_clause = ', '.join(f'{k} = ?' for k in updates)
        db.execute(
            f'UPDATE incidencias_operativas SET {set_clause} WHERE id = ?',
            list(updates.values()) + [inc_id]
        )
        db.commit()
    return jsonify({'ok': True})


@app.delete('/api/incidencias/<int:inc_id>')
@require_auth
def incidencia_delete(inc_id):
    if g.rol not in ('admin', 'operador'):
        return jsonify({'ok': False, 'motivo': 'sin_permiso'}), 403
    db = get_db()
    db.execute('DELETE FROM incidencias_operativas WHERE id = ?', (inc_id,))
    db.commit()
    return jsonify({'ok': True})


@app.get('/api/incidencias/stock')
@require_auth
def incidencias_stock():
    """Incidencias de stock (MAL_STOCK/NO_ENCONTRADO) desde eventos_estado con filtros."""
    db      = get_db()
    desde   = request.args.get('desde', '')
    hasta   = request.args.get('hasta', '')
    sucursal = request.args.get('sucursal', '')

    conds  = ["e.estado_nuevo IN ('MAL_STOCK','NO_ENCONTRADO')"]
    params = []
    if desde:
        conds.append('e.created_at >= ?'); params.append(desde)
    if hasta:
        conds.append('e.created_at <= ?'); params.append(hasta + 'T23:59:59')
    if sucursal:
        conds.append('e.nodo_asignado = ?'); params.append(sucursal)

    rows = db.execute(
        f'''SELECT e.id, e.estado_nuevo,
                   COALESCE(e.nodo_asignado, l.nodo_asignado) as nodo_asignado,
                   e.created_at,
                   l.nro_pedido, l.punto_retiro, l.producto, l.gtin,
                   l.unidades, l.alternativa_nodo, l.notas, l.cruce_id
            FROM eventos_estado e
            JOIN lineas l ON l.id = e.linea_id
            WHERE {" AND ".join(conds)}
            ORDER BY e.created_at DESC''',
        params
    ).fetchall()

    sucursales = db.execute(
        '''SELECT DISTINCT nodo_asignado FROM eventos_estado
           WHERE estado_nuevo IN ('MAL_STOCK','NO_ENCONTRADO')
             AND nodo_asignado IS NOT NULL AND nodo_asignado != ''
           ORDER BY nodo_asignado'''
    ).fetchall()

    return jsonify({
        'ok': True,
        'incidencias': [dict(r) for r in rows],
        'sucursales': [r['nodo_asignado'] for r in sucursales],
    })


@app.get('/api/incidencias/stock/farmacia/<path:nombre>')
@require_auth
def incidencias_stock_farmacia(nombre):
    """Detalle de productos con MAL_STOCK/NO_ENCONTRADO para una sucursal."""
    db     = get_db()
    desde  = request.args.get('desde', '')
    hasta  = request.args.get('hasta', '')

    conds  = ["e.estado_nuevo IN ('MAL_STOCK','NO_ENCONTRADO')", 'e.nodo_asignado = ?']
    params = [nombre]
    if desde:
        conds.append('e.created_at >= ?'); params.append(desde)
    if hasta:
        conds.append('e.created_at <= ?'); params.append(hasta + 'T23:59:59')

    rows = db.execute(
        f'''SELECT e.estado_nuevo,
                   COALESCE(e.nodo_asignado, l.nodo_asignado) as nodo_asignado,
                   e.created_at,
                   l.nro_pedido, l.punto_retiro, l.producto, l.gtin,
                   l.unidades, l.alternativa_nodo, l.notas
            FROM eventos_estado e
            JOIN lineas l ON l.id = e.linea_id
            WHERE {" AND ".join(conds)}
            ORDER BY e.created_at DESC''',
        params
    ).fetchall()
    return jsonify({'ok': True, 'productos': [dict(r) for r in rows], 'farmacia': nombre})


# ── Procesamiento ───────────────────────────────────────────────────────────────

def _procesar_cruce(path_pedidos, path_stock, out_dir):
    """Orquesta el pipeline de procesamiento usando los módulos src/."""
    from src.loader     import cargar_archivo
    from src.matcher    import mapear_columnas_pedidos, mapear_columnas_stock
    from src.services.normalizacion import normalizar_pedidos, normalizar_stock
    from src.services.matching      import ejecutar_matching
    from src.optimizer  import construir_planilla
    from src.exporter   import exportar_excel_profesional, DatosExport

    with open(str(CONFIG_PATH), 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    df_pedidos = cargar_archivo(path_pedidos)
    df_stock   = cargar_archivo(path_stock)

    mapa_pedidos = mapear_columnas_pedidos(df_pedidos, config)
    mapa_stock   = mapear_columnas_stock(df_stock, config)

    df_pedidos = normalizar_pedidos(df_pedidos, mapa_pedidos)
    df_stock   = normalizar_stock(df_stock, mapa_stock)

    resultado_matching = ejecutar_matching(
        df_pedidos, df_stock, mapa_pedidos, mapa_stock, config
    )

    df_ruta, df_sin_stock, _ = construir_planilla(
        df_pedidos=df_pedidos, df_stock=df_stock,
        mapa_pedidos=mapa_pedidos, mapa_stock=mapa_stock, cfg=config,
    )

    excel_path = str(Path(out_dir) / 'planilla_cruce.xlsx')
    datos = DatosExport(
        df_ruta=df_ruta,
        df_sin_stock=df_sin_stock,
        estados_busqueda=config.get('estados_busqueda', []),
        resultado_matching=resultado_matching,
        archivo_pedidos=Path(path_pedidos).name,
        archivo_stock=Path(path_stock).name,
    )
    exportar_excel_profesional(datos, excel_path)

    # Índice GTIN/SKU → LineaMatching para calcular alternativas
    import pandas as _pd
    _col_nodo = mapa_stock["nodo"]
    _col_stk  = mapa_stock["stock"]
    _gtin_idx: dict = {}
    _sku_idx:  dict = {}
    for _lm in resultado_matching.lineas:
        if _lm.gtin and _lm.df_stock is not None:
            _gtin_idx[str(_lm.gtin).strip()] = _lm
        if _lm.sku and _lm.df_stock is not None:
            _sku_idx[str(_lm.sku).strip()] = _lm

    def _get_alternativas(gtin, sku, nodo_asignado):
        lm = _gtin_idx.get(str(gtin).strip()) or _sku_idx.get(str(sku).strip())
        if not lm or lm.df_stock is None:
            return []
        df_tmp = _pd.DataFrame({
            'nodo':  lm.df_stock[_col_nodo].values,
            'stock': _pd.to_numeric(lm.df_stock[_col_stk], errors='coerce').fillna(0).values,
        })
        stock_por_nodo = df_tmp.groupby('nodo')['stock'].sum().sort_values(ascending=False)
        return [
            {'nodo': n, 'stock': int(s)}
            for n, s in stock_por_nodo.items()
            if int(s) > 0 and n != nodo_asignado
        ][:3]

    # Convertir df_ruta al formato de líneas que espera la API
    lineas = []
    if not df_ruta.empty:
        for _, row in df_ruta.iterrows():
            farmacia = row.get('Farmacia', '')
            es_sin_cob = (not farmacia) or farmacia == '— SIN COBERTURA —'
            gtin = str(row.get('GTIN', '') or '')
            sku  = str(row.get('Zetti (ID)', '') or '')
            nodo = None if es_sin_cob else farmacia
            lineas.append({
                'nro_pedido':      str(row.get('N° Pedido', '') or ''),
                'gtin':            gtin,
                'sku':             sku,
                'producto':        str(row.get('Producto', '') or ''),
                'marca':           '',
                'unidades':        int(row.get('Unidades a buscar') or row.get('Cantidad pedida') or 1),
                'nodo_asignado':    nodo,
                'zona':             int(row.get('prioridad', 2)),
                'tier':             int(row.get('_tier', 2)),
                'stock_disponible': int(row.get('Stock sucursal', 0) or 0),
                'alternativas':     _get_alternativas(gtin, sku, nodo),
                'punto_retiro':     str(row.get('Punto de Retiro', '') or ''),
                'cliente_nombre':   str(row.get('Cliente Nombre', '') or ''),
                'cliente_dni':      str(row.get('Cliente DNI', '') or ''),
                'cliente_telefono': str(row.get('Cliente Teléfono', '') or ''),
            })

    # Líneas sin cobertura del df_sin_stock
    if not df_sin_stock.empty:
        for _, row in df_sin_stock.iterrows():
            gtin = str(row.get('GTIN', '') or '')
            sku  = str(row.get('SKU', '') or '')
            lineas.append({
                'nro_pedido':      str(row.get('N° Pedido', '') or ''),
                'gtin':            gtin,
                'sku':             sku,
                'producto':        str(row.get('Producto', '') or ''),
                'marca':           '',
                'unidades':        int(row.get('Unidades', 1) or 1),
                'nodo_asignado':    None,
                'zona':             99,
                'tier':             99,
                'stock_disponible': 0,
                'alternativas':     _get_alternativas(gtin, sku, None),
                'punto_retiro':     str(row.get('Punto de Retiro', '') or ''),
                'cliente_nombre':   str(row.get('Cliente Nombre', '') or ''),
                'cliente_dni':      str(row.get('Cliente DNI', '') or ''),
                'cliente_telefono': str(row.get('Cliente Teléfono', '') or ''),
            })

    con_cob = sum(1 for l in lineas if l['nodo_asignado'])
    sin_cob = sum(1 for l in lineas if not l['nodo_asignado'])
    total   = len(lineas)
    pct     = round(con_cob / total * 100, 1) if total else 0

    return {
        'total':         total,
        'con_cobertura': con_cob,
        'sin_cobertura': sin_cob,
        'pct_cobertura': pct,
        'excel_path':    excel_path,
        'lineas':        lineas,
    }

# ── Gestión de usuarios (admin) ─────────────────────────────────────────────────

@app.get('/api/usuarios')
@require_admin
def usuarios_list():
    db   = get_db()
    rows = db.execute(
        'SELECT id, usuario, nombre, rol, created_at FROM usuarios ORDER BY id'
    ).fetchall()
    return jsonify({'ok': True, 'usuarios': [dict(r) for r in rows]})


@app.post('/api/usuarios')
@require_admin
def usuarios_crear():
    data     = request.get_json(silent=True) or {}
    usuario  = data.get('usuario', '').strip()
    nombre   = data.get('nombre', '').strip()
    password = data.get('password', '')
    rol      = data.get('rol', 'operador')

    if not usuario or not password:
        return jsonify({'ok': False, 'motivo': 'campos_requeridos'}), 400
    if rol not in ('admin', 'operador', 'cadete'):
        return jsonify({'ok': False, 'motivo': 'rol_invalido'}), 400

    db = get_db()
    if db.execute('SELECT id FROM usuarios WHERE usuario = ?', (usuario,)).fetchone():
        return jsonify({'ok': False, 'motivo': 'usuario_existente'}), 409

    db.execute(
        'INSERT INTO usuarios (usuario, nombre, password_hash, rol) VALUES (?, ?, ?, ?)',
        (usuario, nombre, generate_password_hash(password), rol)
    )
    db.commit()
    return jsonify({'ok': True})


@app.patch('/api/usuarios/<int:uid>')
@require_admin
def usuarios_editar(uid):
    data     = request.get_json(silent=True) or {}
    nombre   = data.get('nombre')
    password = data.get('password')
    rol      = data.get('rol')

    db = get_db()
    if not db.execute('SELECT id FROM usuarios WHERE id = ?', (uid,)).fetchone():
        return jsonify({'ok': False, 'motivo': 'no_existe'}), 404

    if nombre is not None:
        db.execute('UPDATE usuarios SET nombre = ? WHERE id = ?', (nombre.strip(), uid))
    if password:
        db.execute('UPDATE usuarios SET password_hash = ? WHERE id = ?',
                   (generate_password_hash(password), uid))
    if rol and rol in ('admin', 'operador', 'cadete'):
        db.execute('UPDATE usuarios SET rol = ? WHERE id = ?', (rol, uid))

    db.commit()
    return jsonify({'ok': True})


@app.delete('/api/usuarios/<int:uid>')
@require_admin
def usuarios_eliminar(uid):
    if uid == g.user_id:
        return jsonify({'ok': False, 'motivo': 'no_eliminar_propio'}), 400
    db = get_db()
    db.execute('DELETE FROM usuarios WHERE id = ?', (uid,))
    db.commit()
    return jsonify({'ok': True})


# ── Configuración del sistema ───────────────────────────────────────────────────

@app.post('/api/sucursales-fijas')
@require_auth
def sucursal_fija_add():
    if g.rol not in ('admin', 'operador'):
        return jsonify({'ok': False, 'motivo': 'Sin permiso'}), 403
    data = request.get_json(force=True)
    nombre = data.get('nombre', '').strip()
    if not nombre:
        return jsonify({'ok': False, 'motivo': 'Nombre requerido'}), 400
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f)
        fijas = cfg.get('sucursales_fijas', [])
        if nombre not in fijas:
            fijas.append(nombre)
            cfg['sucursales_fijas'] = fijas
            with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
                yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    except Exception as e:
        return jsonify({'ok': False, 'motivo': str(e)}), 500
    return jsonify({'ok': True, 'sucursales_fijas': fijas})


@app.delete('/api/sucursales-fijas')
@require_auth
def sucursal_fija_remove():
    if g.rol not in ('admin', 'operador'):
        return jsonify({'ok': False, 'motivo': 'Sin permiso'}), 403
    data = request.get_json(force=True)
    nombre = data.get('nombre', '').strip()
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f)
        fijas = [s for s in cfg.get('sucursales_fijas', []) if s != nombre]
        cfg['sucursales_fijas'] = fijas
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    except Exception as e:
        return jsonify({'ok': False, 'motivo': str(e)}), 500
    return jsonify({'ok': True, 'sucursales_fijas': fijas})


@app.get('/api/config')
@require_admin
def config_leer():
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f)
    except Exception as e:
        return jsonify({'ok': False, 'motivo': str(e)}), 500
    return jsonify({
        'ok': True,
        'asignacion':      cfg.get('asignacion',      {}),
        'optimizacion':    cfg.get('optimizacion',    {}),
        'sucursales_fijas': cfg.get('sucursales_fijas', []),
    })


@app.put('/api/config')
@require_admin
def config_guardar():
    body = request.get_json(force=True)
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f)

        if 'asignacion' in body:
            a = cfg.setdefault('asignacion', {})
            for k in ('nodos_tier_0', 'nodos_tier_1', 'palabras_clave_tier_1',
                      'palabras_clave_tier_3', 'consolidacion_por_pedido',
                      'max_sucursales_por_producto'):
                if k in body['asignacion']:
                    a[k] = body['asignacion'][k]

        if 'optimizacion' in body:
            o = cfg.setdefault('optimizacion', {})
            for k in ('stock_seguridad', 'max_sucursales_por_producto',
                      'stock_sospechoso_umbral', 'max_opciones_override'):
                if k in body['optimizacion']:
                    o[k] = body['optimizacion'][k]

        if 'sucursales_fijas' in body:
            cfg['sucursales_fijas'] = [
                s.strip() for s in body['sucursales_fijas']
                if isinstance(s, str) and s.strip()
            ]

        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    except Exception as e:
        return jsonify({'ok': False, 'motivo': str(e)}), 500
    return jsonify({'ok': True})


@app.get('/api/config/sugerencias')
@require_admin
def config_sugerencias():
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f)
    except Exception as e:
        return jsonify({'ok': False, 'motivo': str(e)}), 500

    asig     = cfg.get('asignacion', {})
    tier0    = {n.lower() for n in asig.get('nodos_tier_0', [])}
    tier1_n  = {n.lower() for n in asig.get('nodos_tier_1', [])}
    kw_tier1 = [k.lower() for k in asig.get('palabras_clave_tier_1', [])]
    kw_tier3 = [k.lower() for k in asig.get('palabras_clave_tier_3', [])]

    def tier_de(nodo):
        nl = nodo.lower()
        if nl in tier0: return 0
        if nl in tier1_n or any(kw in nl for kw in kw_tier1): return 1
        if any(kw in nl for kw in kw_tier3): return 3
        return 2

    db = get_db()

    # ── Estadísticas por nodo ────────────────────────────────────
    rows = db.execute('''
        SELECT nodo_asignado,
               COUNT(*)                                                      AS total,
               COUNT(DISTINCT cruce_id)                                      AS cruces,
               SUM(CASE WHEN estado = 'MAL_STOCK'    THEN 1 ELSE 0 END)     AS mal_stock,
               SUM(CASE WHEN estado = 'ENCONTRADO'   THEN 1 ELSE 0 END)     AS encontrado,
               SUM(CASE WHEN estado = 'NO_ENCONTRADO' THEN 1 ELSE 0 END)    AS no_encontrado
        FROM lineas
        WHERE nodo_asignado IS NOT NULL AND nodo_asignado != ''
          AND estado NOT IN ('PENDIENTE', 'SIN_COBERTURA')
        GROUP BY nodo_asignado
        HAVING total >= 5
        ORDER BY total DESC
    ''').fetchall()

    sugerencias = []

    for r in rows:
        nodo     = r['nodo_asignado']
        total    = r['total']
        enc      = r['encontrado']
        mal      = r['mal_stock']
        no_enc   = r['no_encontrado']
        cruces   = r['cruces']
        resuelto = enc + mal + no_enc
        if resuelto == 0:
            continue

        conf     = enc / resuelto
        mal_rate = mal / resuelto
        tier_act = tier_de(nodo)

        # Promover a Tier 0 — muy confiable y no está ya ahí
        if tier_act > 0 and conf >= 0.95 and cruces >= 3:
            sugerencias.append({
                'id':          f'promover_tier0_{nodo}',
                'tipo':        'promover_tier0',
                'nodo':        nodo,
                'tier_actual': tier_act,
                'confiabilidad': round(conf * 100, 1),
                'mal_stock_pct': round(mal_rate * 100, 1),
                'lineas':      total,
                'cruces':      cruces,
            })
        # Promover a Tier 1 — confiable pero en tier 2/3
        elif tier_act >= 2 and conf >= 0.85 and cruces >= 3:
            sugerencias.append({
                'id':          f'promover_tier1_{nodo}',
                'tipo':        'promover_tier1',
                'nodo':        nodo,
                'tier_actual': tier_act,
                'confiabilidad': round(conf * 100, 1),
                'mal_stock_pct': round(mal_rate * 100, 1),
                'lineas':      total,
                'cruces':      cruces,
            })

        # Bajar tier — mal stock frecuente en sucursal de alta prioridad
        if tier_act <= 1 and mal_rate >= 0.30 and cruces >= 3:
            sugerencias.append({
                'id':          f'bajar_tier_{nodo}',
                'tipo':        'bajar_tier',
                'nodo':        nodo,
                'tier_actual': tier_act,
                'confiabilidad': round(conf * 100, 1),
                'mal_stock_pct': round(mal_rate * 100, 1),
                'lineas':      total,
                'cruces':      cruces,
            })

    # ── Productos sin cobertura recurrente ───────────────────────
    prod_rows = db.execute('''
        SELECT gtin, producto,
               COUNT(*)                                                   AS total,
               SUM(CASE WHEN estado = 'SIN_COBERTURA' THEN 1 ELSE 0 END) AS sin_cob,
               COUNT(DISTINCT cruce_id)                                   AS cruces
        FROM lineas
        WHERE gtin IS NOT NULL AND gtin != ''
        GROUP BY gtin
        HAVING total >= 5 AND CAST(sin_cob AS REAL) / total >= 0.70
        ORDER BY sin_cob DESC
        LIMIT 8
    ''').fetchall()

    for r in prod_rows:
        tasa = r['sin_cob'] / r['total']
        sugerencias.append({
            'id':      f'sin_cobertura_{r["gtin"]}',
            'tipo':    'sin_cobertura_producto',
            'gtin':    r['gtin'],
            'producto': r['producto'] or r['gtin'],
            'tasa':    round(tasa * 100, 1),
            'lineas':  r['total'],
            'cruces':  r['cruces'],
        })

    # ── Stock de seguridad global ────────────────────────────────
    opt       = cfg.get('optimizacion', {})
    stock_seg = opt.get('stock_seguridad', 0)
    t_asig = db.execute(
        "SELECT COUNT(*) FROM lineas WHERE nodo_asignado != '' AND estado NOT IN ('PENDIENTE','SIN_COBERTURA')"
    ).fetchone()[0]
    t_mal  = db.execute(
        "SELECT COUNT(*) FROM lineas WHERE nodo_asignado != '' AND estado = 'MAL_STOCK'"
    ).fetchone()[0]

    if t_asig >= 20 and t_mal / t_asig >= 0.15 and stock_seg < 5:
        sugerencias.append({
            'id':             'stock_seguridad_bajo',
            'tipo':           'stock_seguridad',
            'valor_actual':   stock_seg,
            'valor_sugerido': min(stock_seg + 2, 10),
            'mal_stock_pct':  round(t_mal / t_asig * 100, 1),
        })

    total_cruces = db.execute('SELECT COUNT(*) FROM cruces').fetchone()[0]
    return jsonify({'ok': True, 'sugerencias': sugerencias, 'total_cruces': total_cruces})


# ── Líneas manuales ─────────────────────────────────────────────────────────────

@app.get('/api/cruces/<int:cid>/nodos')
@require_auth
def cruce_nodos(cid):
    db   = get_db()
    rows = db.execute('''
        SELECT DISTINCT nodo_asignado FROM lineas
        WHERE cruce_id = ? AND nodo_asignado IS NOT NULL AND nodo_asignado != ''
        ORDER BY nodo_asignado
    ''', (cid,)).fetchall()
    return jsonify({'ok': True, 'nodos': [r['nodo_asignado'] for r in rows]})


@app.post('/api/cruces/<int:cid>/lineas')
@require_auth
def cruce_agregar_linea(cid):
    if g.rol not in ('admin', 'operador'):
        return jsonify({'ok': False, 'motivo': 'sin_permiso'}), 403

    body     = request.get_json(force=True)
    gtin     = (body.get('gtin')          or '').strip()
    sku      = (body.get('sku')           or '').strip()
    producto = (body.get('producto')      or '').strip()
    nodo     = (body.get('nodo_asignado') or '').strip()
    unidades = int(body.get('unidades')   or 1)
    nro_ped  = (body.get('nro_pedido')    or '').strip()
    notas    = (body.get('notas')         or '').strip()

    if not gtin and not sku:
        return jsonify({'ok': False, 'motivo': 'se_requiere_gtin_o_sku'}), 400

    db    = get_db()
    cruce = db.execute('SELECT id FROM cruces WHERE id = ?', (cid,)).fetchone()
    if not cruce:
        return jsonify({'ok': False, 'motivo': 'cruce_no_encontrado'}), 404

    # Calcular tier del nodo si fue especificado
    tier = 2
    if nodo:
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                cfg = yaml.safe_load(f)
            asig    = cfg.get('asignacion', {})
            tier0   = {n.lower() for n in asig.get('nodos_tier_0', [])}
            tier1_n = {n.lower() for n in asig.get('nodos_tier_1', [])}
            kw1     = [k.lower() for k in asig.get('palabras_clave_tier_1', [])]
            kw3     = [k.lower() for k in asig.get('palabras_clave_tier_3', [])]
            nl = nodo.lower()
            if nl in tier0:                               tier = 0
            elif nl in tier1_n or any(k in nl for k in kw1): tier = 1
            elif any(k in nl for k in kw3):               tier = 3
        except Exception:
            pass

    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute('''
        INSERT INTO lineas
            (cruce_id, nro_pedido, gtin, sku, producto, unidades,
             nodo_asignado, tier, estado, notas, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (cid, nro_ped, gtin, sku, producto, unidades,
          nodo or None, tier,
          'PENDIENTE' if nodo else 'SIN_COBERTURA',
          notas, now))

    db.execute('UPDATE cruces SET total_lineas = total_lineas + 1 WHERE id = ?', (cid,))
    db.commit()
    return jsonify({'ok': True, 'linea_id': cur.lastrowid})


# ── Jornada del día ─────────────────────────────────────────────────────────────

@app.get('/api/jornada/hoy')
@require_auth
def jornada_hoy():
    fecha = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    db    = get_db()
    row   = db.execute(
        'SELECT puntos, cruce_id FROM jornadas WHERE fecha = ? AND cadete_id = ?',
        (fecha, g.user_id)
    ).fetchone()
    if not row:
        return jsonify({'ok': True, 'jornada': None})
    return jsonify({'ok': True, 'jornada': {
        'puntos':   json.loads(row['puntos']),
        'cruce_id': row['cruce_id'],
    }})


@app.post('/api/jornada')
@require_auth
def jornada_guardar():
    body     = request.get_json(force=True)
    cruce_id = body.get('cruce_id')
    puntos   = body.get('puntos', [])
    fecha    = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    db = get_db()
    db.execute('''
        INSERT INTO jornadas (fecha, cadete_id, cruce_id, puntos)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(fecha, cadete_id) DO UPDATE SET
            puntos   = excluded.puntos,
            cruce_id = excluded.cruce_id
    ''', (fecha, g.user_id, cruce_id, json.dumps(puntos, ensure_ascii=False)))
    db.commit()
    return jsonify({'ok': True})


@app.get('/api/cruces/<int:cid>/jornadas')
@require_auth
def cruce_jornadas(cid):
    fecha = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    db    = get_db()
    rows  = db.execute('''
        SELECT j.cadete_id, u.nombre, u.usuario, j.puntos
        FROM jornadas j
        JOIN usuarios u ON u.id = j.cadete_id
        WHERE j.fecha = ? AND j.cruce_id = ?
    ''', (fecha, cid)).fetchall()
    return jsonify({'ok': True, 'jornadas': [
        {
            'cadete_id': r['cadete_id'],
            'nombre':    r['nombre'] or r['usuario'],
            'puntos':    json.loads(r['puntos']),
        }
        for r in rows
    ]})


# ── Tiempos por farmacia ────────────────────────────────────────────────────────

@app.get('/api/analytics/tiempos-farmacia')
@require_auth
def analytics_tiempos_farmacia():
    from datetime import datetime as dt
    db   = get_db()
    rows = db.execute('''
        SELECT
            nodo_asignado,
            cruce_id,
            MIN(CASE WHEN estado NOT IN ('PENDIENTE','SIN_COBERTURA') THEN updated_at END) AS primera_accion,
            MAX(CASE WHEN estado NOT IN ('PENDIENTE','SIN_COBERTURA') THEN updated_at END) AS ultima_accion
        FROM lineas
        WHERE nodo_asignado IS NOT NULL AND nodo_asignado != ''
          AND updated_at IS NOT NULL
        GROUP BY cruce_id, nodo_asignado
        HAVING primera_accion IS NOT NULL
    ''').fetchall()

    farm_muestras = {}
    for r in rows:
        nodo = r['nodo_asignado']
        try:
            t1  = dt.fromisoformat(r['primera_accion'].replace('Z', '+00:00'))
            t2  = dt.fromisoformat(r['ultima_accion'].replace('Z', '+00:00'))
            dur = (t2 - t1).total_seconds() / 60
            if dur < 1:
                continue
        except Exception:
            continue
        farm_muestras.setdefault(nodo, []).append({
            'cruce_id':    r['cruce_id'],
            'duracion_min': round(dur, 1),
            'fecha':        r['primera_accion'],
        })

    result = []
    for nodo, muestras in farm_muestras.items():
        muestras.sort(key=lambda x: x['fecha'])
        durs     = [m['duracion_min'] for m in muestras]
        promedio = sum(durs) / len(durs)
        ultima   = durs[-1]
        if len(durs) >= 3:
            hist_prom = sum(durs[:-1]) / len(durs[:-1])
            tendencia = ('mejorando'  if ultima < hist_prom * 0.85 else
                         'empeorando' if ultima > hist_prom * 1.15 else 'estable')
        else:
            tendencia = 'sin_datos'
        result.append({
            'nodo':         nodo,
            'promedio_min': round(promedio, 1),
            'min_min':      round(min(durs), 1),
            'max_min':      round(max(durs), 1),
            'ultima_min':   round(ultima, 1),
            'muestras':     len(muestras),
            'tendencia':    tendencia,
        })

    result.sort(key=lambda x: x['promedio_min'], reverse=True)
    return jsonify({'ok': True, 'farmacias': result})


@app.get('/api/cruces/<int:cid>/timing')
@require_auth
def cruce_timing(cid):
    from datetime import datetime as dt
    db   = get_db()
    rows = db.execute('''
        SELECT
            nodo_asignado,
            MIN(CASE WHEN estado NOT IN ('PENDIENTE','SIN_COBERTURA') THEN updated_at END) AS primera_accion,
            MAX(CASE WHEN estado NOT IN ('PENDIENTE','SIN_COBERTURA') THEN updated_at END) AS ultima_accion,
            SUM(CASE WHEN estado = 'PENDIENTE' THEN 1 ELSE 0 END) AS pendientes,
            COUNT(*) AS total
        FROM lineas
        WHERE cruce_id = ?
          AND nodo_asignado IS NOT NULL AND nodo_asignado != ''
        GROUP BY nodo_asignado
    ''', (cid,)).fetchall()

    ahora     = datetime.now(timezone.utc)
    farmacias = {}
    for r in rows:
        nodo    = r['nodo_asignado']
        primera = r['primera_accion']
        ultima  = r['ultima_accion']
        pend    = r['pendientes']
        dur     = None

        if primera is None:
            estado = 'pendiente'
        elif pend == 0:
            estado = 'completada'
            try:
                t1  = dt.fromisoformat(primera.replace('Z', '+00:00'))
                t2  = dt.fromisoformat(ultima.replace('Z', '+00:00'))
                dur = round((t2 - t1).total_seconds() / 60, 1)
            except Exception:
                pass
        else:
            estado = 'en_curso'
            try:
                t1  = dt.fromisoformat(primera.replace('Z', '+00:00'))
                dur = round((ahora - t1).total_seconds() / 60, 1)
            except Exception:
                pass

        farmacias[nodo] = {
            'primera_accion': primera,
            'ultima_accion':  ultima,
            'pendientes':     pend,
            'total':          r['total'],
            'estado':         estado,
            'duracion_min':   dur,
        }

    return jsonify({'ok': True, 'farmacias': farmacias})


# ── Health check ────────────────────────────────────────────────────────────────

@app.get('/api/health')
def health():
    return jsonify({'ok': True, 'status': 'up'})


# ── Servir frontend ─────────────────────────────────────────────────────────────

@app.get('/')
def root():
    resp = send_from_directory(str(BASE_DIR), 'login.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

@app.get('/<path:path>')
def static_files(path):
    resp = send_from_directory(str(BASE_DIR), path)
    if path.endswith('.html'):
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

# ── Main ────────────────────────────────────────────────────────────────────────

with app.app_context():
    init_db()

if __name__ == '__main__':
    print('\n  Cruce Stock Farmacias')
    print('  http://localhost:5050\n')
    app.run(host='0.0.0.0', port=5050, debug=True)
