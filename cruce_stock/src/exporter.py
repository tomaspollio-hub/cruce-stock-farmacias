"""
exporter.py
Genera el Excel final con:
  - Pestaña "Planilla Cadete"  : la ruta de búsqueda lista para usar
  - Pestaña "Sin Cobertura"    : productos sin stock en ninguna sucursal
  - Pestaña "Log"              : registro de errores y advertencias

Características:
  - Dropdown en columna "Estado de búsqueda" con openpyxl DataValidation
  - Formato condicional: verde = encontrado, rojo = sin cobertura, amarillo = llamar
  - Encabezados en negrita con color de fondo
  - Columnas con ancho ajustado automáticamente
"""

from __future__ import annotations
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.logger import get_logger, get_log_records

logger = get_logger(__name__)

# ── Paleta de colores ────────────────────────────────────────
COLOR_HEADER_RUTA    = "2E4057"   # azul oscuro
COLOR_HEADER_SIN     = "8B0000"   # rojo oscuro
COLOR_HEADER_LOG     = "4A4A4A"   # gris oscuro
COLOR_BUSQUEDA       = "FFFFFF"
COLOR_ENCONTRADO     = "C6EFCE"   # verde claro
COLOR_MAL_STOCK      = "FFEB9C"   # amarillo
COLOR_LLAMAR         = "BDD7EE"   # celeste
COLOR_SIN_COBERTURA  = "FFC7CE"   # rojo claro
COLOR_LLAMAR_CLIENTE = "FFC7CE"


def _aplicar_encabezado(ws, color_hex: str):
    """Pone negrita, fondo de color y texto blanco en la primera fila."""
    fill = PatternFill("solid", fgColor=color_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_ancho(ws):
    """Ajusta el ancho de cada columna al contenido más largo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _df_to_sheet(ws, df: pd.DataFrame):
    """Escribe un DataFrame en un worksheet (incluye encabezados)."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def _agregar_dropdown_estado(ws, estados: list[str], col_letra: str, n_filas: int):
    """
    Agrega validación de datos (dropdown) en la columna de estado
    desde la fila 2 hasta n_filas+1.
    """
    formula = '"' + ",".join(estados) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False,
        showDropDown=False,  # False = mostrar el icono de dropdown
    )
    dv.sqref = f"{col_letra}2:{col_letra}{n_filas + 1}"
    ws.add_data_validation(dv)


def _agregar_formato_condicional_estado(ws, col_letra: str, n_filas: int):
    """
    Colorea la fila completa según el valor de la columna Estado.
    Usa FormulaRule para aplicar color a toda la fila.
    """
    ultima_col = get_column_letter(ws.max_column)
    rango = f"A2:{ultima_col}{n_filas + 1}"

    reglas = [
        ("Encontrado",           COLOR_ENCONTRADO),
        ("Mal stock",            COLOR_MAL_STOCK),
        ("Mal stock - Resuelto", COLOR_MAL_STOCK),
        ("Llamar a suc",         COLOR_LLAMAR),
        ("Llamar cliente",       COLOR_LLAMAR_CLIENTE),
        ("— SIN COBERTURA —",    COLOR_SIN_COBERTURA),
    ]

    for valor, color in reglas:
        fill = PatternFill("solid", fgColor=color)
        # La fórmula ancla a la columna de estado (col_letra) pero aplica a la fila
        formula = f'${col_letra}2="{valor}"'
        ws.conditional_formatting.add(
            rango,
            FormulaRule(formula=[formula], fill=fill)
        )


def exportar_excel(
    path_salida: str,
    df_ruta: pd.DataFrame,
    df_sin_stock: pd.DataFrame,
    estados_busqueda: list[str],
):
    """
    Genera el archivo Excel final.

    path_salida        : ruta completa del archivo a crear
    df_ruta            : planilla del cadete
    df_sin_stock       : productos sin cobertura
    estados_busqueda   : lista de opciones para el dropdown
    """
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    # ── Pestaña 1: Planilla Cadete ───────────────────────────
    ws_ruta = wb.create_sheet("Planilla Cadete")

    if df_ruta.empty:
        ws_ruta.append(["Sin datos para mostrar"])
    else:
        _df_to_sheet(ws_ruta, df_ruta)
        _aplicar_encabezado(ws_ruta, COLOR_HEADER_RUTA)
        _ajustar_ancho(ws_ruta)

        # Detectar columna "Estado de búsqueda"
        headers = [cell.value for cell in ws_ruta[1]]
        if "Estado de búsqueda" in headers:
            col_idx  = headers.index("Estado de búsqueda") + 1
            col_letra = get_column_letter(col_idx)
            n_filas  = len(df_ruta)
            _agregar_dropdown_estado(ws_ruta, estados_busqueda, col_letra, n_filas)
            _agregar_formato_condicional_estado(ws_ruta, col_letra, n_filas)

        # Congelar la primera fila
        ws_ruta.freeze_panes = "A2"

    # ── Pestaña 2: Sin Cobertura ─────────────────────────────
    ws_sin = wb.create_sheet("Sin Cobertura")
    if df_sin_stock.empty:
        ws_sin.append(["✅ Todos los productos tienen cobertura en sucursales"])
    else:
        _df_to_sheet(ws_sin, df_sin_stock)
        _aplicar_encabezado(ws_sin, COLOR_HEADER_SIN)
        _ajustar_ancho(ws_sin)
        ws_sin.freeze_panes = "A2"

    # ── Pestaña 3: Log ───────────────────────────────────────
    ws_log = wb.create_sheet("Log")
    registros = get_log_records()
    if registros:
        df_log = pd.DataFrame(registros)
        _df_to_sheet(ws_log, df_log)
        _aplicar_encabezado(ws_log, COLOR_HEADER_LOG)
        _ajustar_ancho(ws_log)
    else:
        ws_log.append(["Sin advertencias ni errores registrados"])

    # ── Guardar ──────────────────────────────────────────────
    wb.save(path_salida)
    logger.info(f"Archivo exportado: {path_salida}")
    return path_salida


def generar_nombre_salida(carpeta: str) -> str:
    """Genera un nombre de archivo con timestamp para evitar sobreescribir."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{carpeta}/planilla_cadete_{ts}.xlsx"


# ════════════════════════════════════════════════════════════
#  EXPORTADOR PROFESIONAL — 4 hojas
# ════════════════════════════════════════════════════════════

from dataclasses import dataclass, field as dc_field

# ── Paleta extendida ─────────────────────────────────────────
_C = {
    "navy":        "1E3A5F",   # header principal
    "navy_mid":    "2E5A8E",   # header secundario
    "navy_light":  "EBF2FA",   # fila par (alternado)
    "white":       "FFFFFF",
    "found":       "C6EFCE",   # encontrado
    "found_font":  "276221",
    "missing":     "FFC7CE",   # no encontrado / sin cobertura
    "missing_font":"9C0006",
    "review":      "FFEB9C",   # en revisión / mal stock
    "review_font": "9C6500",
    "reassigned":  "BDD7EE",   # reasignado
    "reassigned_font":"185D8C",
    "pending":     "F8FAFC",   # búsqueda / pendiente (solo fondo suave)
    "resumen_key": "D6E4F7",   # fondo clave en hoja RESUMEN
    "incidencia_warn": "FFF2CC",
    "tab_op":      "2563EB",   # color pestaña OPERATIVO
    "tab_cadete":  "059669",   # color pestaña CADETE
    "tab_inc":     "E11D48",   # color pestaña INCIDENCIAS
    "tab_res":     "1E3A5F",   # color pestaña RESUMEN
}

# Columnas internas que nunca van al Excel
_INTERNAS = {
    "_gtin_key", "prioridad", "_criterio", "_tier",
    "_consolida_pedido", "_gtin_norm", "_sku_norm",
    "_unidades_int", "_id_norm",
}

# Mapa de estado → (color fondo, color fuente)
_ESTADO_COLORES: dict[str, tuple[str, str]] = {
    "Encontrado":            (_C["found"],      _C["found_font"]),
    "Mal stock - Resuelto":  (_C["found"],      _C["found_font"]),
    "No encontrado":         (_C["missing"],    _C["missing_font"]),
    "Sin cobertura":         (_C["missing"],    _C["missing_font"]),
    "Llamar cliente":        (_C["missing"],    _C["missing_font"]),
    "— SIN COBERTURA —":     (_C["missing"],    _C["missing_font"]),
    "Mal stock":             (_C["review"],     _C["review_font"]),
    "Requiere revisión":     (_C["review"],     _C["review_font"]),
    "En revisión":           (_C["review"],     _C["review_font"]),
    "Llamar a suc":          (_C["reassigned"], _C["reassigned_font"]),
    "Reasignado":            (_C["reassigned"], _C["reassigned_font"]),
}


@dataclass
class DatosExport:
    """
    Paquete de datos que recibe exportar_excel_profesional().
    Desacopla el exportador de cualquier estado de UI.
    """
    df_ruta:           "pd.DataFrame"
    df_sin_stock:      "pd.DataFrame"
    estados_busqueda:  list[str]
    gestor_estados:    object        = None   # GestorEstados | None
    resultado_matching: object       = None   # ResultadoMatching | None
    resumenes_pedidos: list          = dc_field(default_factory=list)
    archivo_pedidos:   str           = ""
    archivo_stock:     str           = ""
    pedidos_unicos:    int           = 0


# ── Helpers de formato ───────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _border_thin() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _escribir_hdr(ws, fila: int, valores: list, color_fondo: str, color_texto="FFFFFF"):
    fill = _fill(color_fondo)
    font = _font(bold=True, color=color_texto)
    borde = _border_thin()
    for c, val in enumerate(valores, 1):
        cell = ws.cell(row=fila, column=c, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde


def _escribir_fila(ws, fila: int, valores: list, alternado: bool = False, estado: str = ""):
    """Escribe una fila con color alternado y, si tiene estado, colorea según el estado."""
    bg = _C["navy_light"] if alternado else _C["white"]
    if estado and estado in _ESTADO_COLORES:
        bg, font_color = _ESTADO_COLORES[estado]
    else:
        font_color = "000000"

    fill = _fill(bg)
    borde = _border_thin()
    for c, val in enumerate(valores, 1):
        cell = ws.cell(row=fila, column=c, value=val)
        cell.fill = fill
        cell.font = _font(color=font_color)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = borde


def _ajustar_ancho_pro(ws, min_w=8, max_w=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)


def _agregar_autofilter(ws, n_cols: int):
    """Agrega AutoFilter en la fila 1."""
    ultima = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{ultima}1"


def _colorear_columna_estado(ws, col_idx: int, fila_ini: int, fila_fin: int):
    """Colorea celda a celda la columna de estado (más preciso que formato condicional)."""
    for r in range(fila_ini, fila_fin + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = str(cell.value or "")
        if val in _ESTADO_COLORES:
            bg, fc = _ESTADO_COLORES[val]
            cell.fill = _fill(bg)
            cell.font = _font(bold=True, color=fc)


# ── Hoja 1: RESUMEN ──────────────────────────────────────────

def _hoja_resumen(wb: "Workbook", datos: DatosExport, ts: str):
    ws = wb.create_sheet("RESUMEN")
    ws.sheet_properties.tabColor = _C["tab_res"]
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    def kv(fila, clave, valor, bold_val=False):
        c_key = ws.cell(row=fila, column=1, value=clave)
        c_key.fill = _fill(_C["resumen_key"])
        c_key.font = _font(bold=True, color=_C["navy"])
        c_key.alignment = Alignment(vertical="center")
        c_key.border = _border_thin()
        c_val = ws.cell(row=fila, column=2, value=valor)
        c_val.font = _font(bold=bold_val, color="000000")
        c_val.alignment = Alignment(vertical="center")
        c_val.border = _border_thin()

    def seccion(fila, titulo):
        cell = ws.cell(row=fila, column=1, value=titulo)
        cell.fill = _fill(_C["navy"])
        cell.font = _font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(vertical="center")
        ws.merge_cells(f"A{fila}:B{fila}")
        ws.row_dimensions[fila].height = 22

    # Título
    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="Global Ecommerce — Cruce de Stock")
    t.fill = _fill(_C["navy"])
    t.font = _font(bold=True, color="FFFFFF", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    seccion(3, "📅  INFORMACIÓN DEL CRUCE")
    kv(4, "Generado el",       ts)
    kv(5, "Archivo pedidos",   datos.archivo_pedidos or "—")
    kv(6, "Archivo stock",     datos.archivo_stock   or "—")

    # Métricas de cobertura
    df = datos.df_ruta
    df_s = datos.df_sin_stock
    total_filas   = len(df) if df is not None else 0
    sin_cob_ruta  = int((df["Farmacia"] == "— SIN COBERTURA —").sum()) if total_filas else 0
    con_cob_ruta  = total_filas - sin_cob_ruta
    pct_cob       = round(con_cob_ruta / total_filas * 100, 1) if total_filas else 0.0
    pedidos_u     = datos.pedidos_unicos
    suc_set: set[str] = set()
    if total_filas and "Farmacia" in df.columns:
        suc_set = {f for f in df["Farmacia"].unique() if f != "— SIN COBERTURA —"}

    seccion(8, "📊  COBERTURA")
    kv(9,  "Pedidos únicos",           pedidos_u,     bold_val=True)
    kv(10, "Líneas en planilla",       total_filas,   bold_val=True)
    kv(11, "Con cobertura",            con_cob_ruta,  bold_val=True)
    kv(12, "Sin cobertura",            sin_cob_ruta + len(df_s) if df_s is not None else sin_cob_ruta, bold_val=True)
    kv(13, "Sucursales involucradas",  len(suc_set),  bold_val=True)
    kv(14, "% Cobertura",             f"{pct_cob}%",  bold_val=True)

    # Métricas de matching si están disponibles
    rm = datos.resultado_matching
    if rm:
        rs = rm.resumen
        seccion(16, "🔍  MATCHING")
        kv(17, "Total líneas pedido",  rs.total_lineas)
        kv(18, "Match por GTIN",       rs.con_match_gtin)
        kv(19, "Match por SKU",        rs.con_match_sku)
        kv(20, "Sin match",            rs.sin_match)
        kv(21, "Ambiguos",             rs.ambiguos)
        kv(22, "% Cobertura matching", f"{rs.pct_cobertura()}%")

    # Resumen por pedido si está disponible
    rps = datos.resumenes_pedidos
    if rps:
        fila = 24
        seccion(fila, "🗂  RESUMEN POR PEDIDO")
        fila += 1
        _escribir_hdr(ws, fila, ["Pedido", "Sucursales", "Con cobertura", "Sin cobertura"], _C["navy_mid"])
        fila += 1
        for i, rp in enumerate(rps):
            alt = (i % 2 == 1)
            _escribir_fila(ws, fila, [
                rp.nro_pedido,
                rp.total_sucursales,
                rp.filas_con_cobertura,
                rp.filas_sin_cobertura,
            ], alternado=alt)
            fila += 1


# ── Hoja 2: OPERATIVO_POR_PEDIDO ─────────────────────────────

_COLS_OPERATIVO = [
    ("N° Pedido",         "N° Pedido"),
    ("Producto",          "Producto"),
    ("Tipo / Variante",   "Tipo / Variante"),
    ("GTIN",              "GTIN"),
    ("Zetti (ID)",        "Zetti (ID)"),
    ("Cantidad pedida",   "Cantidad"),
    ("Farmacia",          "Sucursal asignada"),
    ("Stock sucursal",    "Stock detectado"),
    ("Unidades a buscar", "Unidades a buscar"),
    ("Zona",              "Zona"),
    ("Estado de búsqueda","Estado"),
    ("⚠️ Stock",          "⚠️ Stock"),
    ("_criterio",         "Criterio asignación"),
    ("Asignación original","Asignación original"),
    ("Observación",       "Observación"),
    ("Reasignado",        "Reasignado"),
]


def _hoja_operativo(wb: "Workbook", datos: DatosExport):
    ws = wb.create_sheet("OPERATIVO_POR_PEDIDO")
    ws.sheet_properties.tabColor = _C["tab_op"]

    df = datos.df_ruta
    if df is None or df.empty:
        ws.append(["Sin datos"]); return

    # Seleccionar y renombrar columnas disponibles
    cols_src = [s for s, _ in _COLS_OPERATIVO if s in df.columns or s in _INTERNAS]
    cols_dst = [d for s, d in _COLS_OPERATIVO if s in df.columns or (s in _INTERNAS and s in df.columns)]
    # Reconstruir respetando las que existen
    pares = [(s, d) for s, d in _COLS_OPERATIVO if s in df.columns]

    hdrs = [d for _, d in pares]
    _escribir_hdr(ws, 1, hdrs, _C["navy"])

    col_estado_idx = next((i + 1 for i, (_, d) in enumerate(pares) if d == "Estado"), None)

    df_sorted = df.sort_values(
        [c for c in ["N° Pedido", "prioridad", "Farmacia"] if c in df.columns]
    ).reset_index(drop=True)

    pedido_actual = None
    for i, (_, row) in enumerate(df_sorted.iterrows()):
        nro = str(row.get("N° Pedido", ""))
        # Fila de separador de pedido
        if nro != pedido_actual and nro:
            pedido_actual = nro
            fila_sep = ws.max_row + 1
            cell = ws.cell(row=fila_sep, column=1, value=f"  Pedido #{nro}")
            cell.fill = _fill(_C["navy_mid"])
            cell.font = _font(bold=True, color="FFFFFF")
            # Combinar visualmente (no merge real para no romper filtros)
            for c in range(2, len(hdrs) + 1):
                ws.cell(row=fila_sep, column=c).fill = _fill(_C["navy_mid"])

        fila_r = ws.max_row + 1
        estado = str(row.get("Estado de búsqueda", ""))
        vals = [row.get(s, "") for s, _ in pares]
        # Limpiar None/NaN
        vals = ["" if (v != v or v is None) else v for v in vals]  # NaN check
        _escribir_fila(ws, fila_r, vals, alternado=(i % 2 == 1), estado=estado)

    ws.freeze_panes = "A2"
    _agregar_autofilter(ws, len(hdrs))
    _ajustar_ancho_pro(ws)

    # Dropdown en columna Estado
    if col_estado_idx:
        col_l = get_column_letter(col_estado_idx)
        _agregar_dropdown_estado(ws, datos.estados_busqueda, col_l, ws.max_row - 1)


# ── Hoja 3: VISTA_CADETE ─────────────────────────────────────

_COLS_CADETE = [
    ("_orden",            "Orden"),
    ("Farmacia",          "Sucursal"),
    ("Zona",              "Zona"),
    ("N° Pedido",         "N° Pedido"),
    ("Producto",          "Producto"),
    ("GTIN",              "GTIN"),
    ("Zetti (ID)",        "Zetti (ID)"),
    ("Unidades a buscar", "Cantidad"),
    ("Estado de búsqueda","Estado"),
    ("Observación",       "Observación"),
    ("⚠️ Stock",          "⚠️ Stock"),
]


def _hoja_cadete(wb: "Workbook", datos: DatosExport):
    ws = wb.create_sheet("VISTA_CADETE")
    ws.sheet_properties.tabColor = _C["tab_cadete"]

    df = datos.df_ruta
    if df is None or df.empty:
        ws.append(["Sin datos"]); return

    # Ordenar: por prioridad (zona), luego farmacia, luego pedido
    sort_cols = [c for c in ["prioridad", "Farmacia", "N° Pedido"] if c in df.columns]
    df_sorted = df.sort_values(sort_cols).reset_index(drop=True)

    # Aplicar estados del gestor si está disponible
    if datos.gestor_estados is not None:
        df_sorted = df_sorted.copy()
        if "Estado de búsqueda" in df_sorted.columns:
            for idx in df_sorted.index:
                orig_idx = df_sorted.index[idx] if hasattr(df_sorted.index, '__getitem__') else idx
                df_sorted.at[idx, "Estado de búsqueda"] = datos.gestor_estados.estado_str(orig_idx)

    # Agregar columna _orden
    df_sorted["_orden"] = range(1, len(df_sorted) + 1)

    pares = [(s, d) for s, d in _COLS_CADETE if s in df_sorted.columns]
    hdrs = [d for _, d in pares]
    _escribir_hdr(ws, 1, hdrs, _C["tab_cadete"])

    farm_actual = None
    for i, (_, row) in enumerate(df_sorted.iterrows()):
        farm = str(row.get("Farmacia", ""))
        # Separador visual por sucursal
        if farm != farm_actual and farm and farm != "— SIN COBERTURA —":
            farm_actual = farm
            zona = str(row.get("Zona", ""))
            fila_sep = ws.max_row + 1
            label = f"  {farm}" + (f"  [{zona}]" if zona else "")
            cell = ws.cell(row=fila_sep, column=1, value=label)
            cell.fill = _fill(_C["tab_cadete"])
            cell.font = _font(bold=True, color="FFFFFF")
            for c in range(2, len(hdrs) + 1):
                ws.cell(row=fila_sep, column=c).fill = _fill(_C["tab_cadete"])

        fila_r = ws.max_row + 1
        estado = str(row.get("Estado de búsqueda", ""))
        vals = [row.get(s, "") for s, _ in pares]
        vals = ["" if (isinstance(v, float) and v != v) else v for v in vals]
        _escribir_fila(ws, fila_r, vals, alternado=(i % 2 == 1), estado=estado)

    ws.freeze_panes = "A2"
    _agregar_autofilter(ws, len(hdrs))
    _ajustar_ancho_pro(ws)


# ── Hoja 4: INCIDENCIAS ──────────────────────────────────────

_ESTADOS_INCIDENCIA = {
    "No encontrado", "Sin cobertura", "En revisión",
    "Requiere revisión", "Mal stock", "Llamar cliente",
    "— SIN COBERTURA —",
}

_COLS_INCIDENCIAS = [
    "Tipo", "N° Pedido", "Producto", "GTIN", "Zetti (ID)",
    "Motivo / Falla", "Sucursal asignada", "Alternativas", "Estado", "Requiere revisión",
]


def _hoja_incidencias(wb: "Workbook", datos: DatosExport):
    ws = wb.create_sheet("INCIDENCIAS")
    ws.sheet_properties.tabColor = _C["tab_inc"]

    _escribir_hdr(ws, 1, _COLS_INCIDENCIAS, _C["tab_inc"])

    filas: list[list] = []

    df = datos.df_ruta
    df_s = datos.df_sin_stock
    gestor = datos.gestor_estados

    # ── Filas del df_ruta con estado problemático ────────────
    if df is not None and not df.empty:
        for idx, row in df.iterrows():
            estado_raw = str(row.get("Estado de búsqueda", ""))
            estado = gestor.estado_str(idx) if gestor else estado_raw
            farmacia = str(row.get("Farmacia", ""))

            es_incidencia = (
                estado in _ESTADOS_INCIDENCIA
                or farmacia == "— SIN COBERTURA —"
            )
            if not es_incidencia:
                continue

            if farmacia == "— SIN COBERTURA —":
                tipo   = "Sin cobertura"
                motivo = "Sin stock en todas las sucursales"
            else:
                tipo   = estado
                motivo = (gestor.ultimo_motivo(idx) if gestor else "") or estado

            alternativas = ""
            if datos.resultado_matching:
                lineas = [
                    l for l in datos.resultado_matching.lineas
                    if l.gtin == str(row.get("GTIN", ""))
                ]
                if lineas and lineas[0].nodos_disponibles:
                    alt_list = [n for n in lineas[0].nodos_disponibles if n != farmacia]
                    alternativas = ", ".join(alt_list[:3])

            req_revision = "Sí" if estado in {"En revisión", "Requiere revisión", "Mal stock"} else ""

            filas.append([
                tipo,
                str(row.get("N° Pedido", "")),
                str(row.get("Producto", "")),
                str(row.get("GTIN", "")),
                str(row.get("Zetti (ID)", "")),
                motivo,
                farmacia,
                alternativas,
                estado,
                req_revision,
            ])

    # ── Filas del df_sin_stock (sin match en sistema) ────────
    if df_s is not None and not df_s.empty:
        for _, row in df_s.iterrows():
            filas.append([
                "Sin match",
                str(row.get("N° Pedido", "")),
                str(row.get("Producto",  "")),
                str(row.get("GTIN",      "")),
                str(row.get("SKU",       "")),
                str(row.get("Motivo",    "Sin match en stock")),
                "—",
                "—",
                "Sin cobertura",
                "Sí",
            ])

    if not filas:
        ws.cell(row=2, column=1, value="✅ Sin incidencias en este cruce")
        return

    for i, vals in enumerate(filas):
        fila_r = i + 2
        estado = vals[8]  # columna "Estado"
        _escribir_fila(ws, fila_r, vals, alternado=(i % 2 == 1), estado=estado)

    ws.freeze_panes = "A2"
    _agregar_autofilter(ws, len(_COLS_INCIDENCIAS))
    _ajustar_ancho_pro(ws)


# ── Entry point ──────────────────────────────────────────────

def exportar_excel_profesional(datos: DatosExport, path_salida: str) -> str:
    """
    Genera el Excel profesional con 4 hojas:
      1. RESUMEN               — métricas del cruce
      2. OPERATIVO_POR_PEDIDO  — planilla completa agrupada por pedido
      3. VISTA_CADETE          — ordenado por sucursal para el cadete
      4. INCIDENCIAS           — problemas y excepciones

    Args:
        datos:       DatosExport con toda la información necesaria.
        path_salida: Ruta del archivo .xlsx a generar.

    Returns:
        path_salida (para encadenar si se necesita).
    """
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_resumen(wb, datos, ts)
    _hoja_operativo(wb, datos)
    _hoja_cadete(wb, datos)
    _hoja_incidencias(wb, datos)

    wb.save(path_salida)
    logger.info(f"Excel profesional exportado: {path_salida} ({len(wb.sheetnames)} hojas)")
    return path_salida
